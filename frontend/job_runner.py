# frontend/job_runner.py
"""
Bridge between Streamlit UI and v1 backend.

Stable design:
- Give every request its own UUID-scoped inputs, output, logs and config.
- Run the backend package as an isolated subprocess.
- Pass selected ATO / ITR policy and owned paths through environment variables.
- Optionally run a Gemini face-check against user inputs + workbook summary.
- Copy the owned output into a session-scoped download history.
- Remove transient job files by default.

This avoids calling backend internals directly and avoids tuple/reports mismatch.
"""

from __future__ import annotations

import json
import logging
import os
import shutil
import subprocess
import sys
import traceback
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, BinaryIO

from tax_calculators.company_tax import assess_base_rate_entity
from tax_calculators.validation import CalculatorError, to_decimal


try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


# ── Paths ─────────────────────────────────────────────────────────────────────

FRONTEND_DIR = Path(__file__).resolve().parent
ROOT_DIR = FRONTEND_DIR.parent
V1_DIR = ROOT_DIR / "v1"

UPLOADS_DIR = FRONTEND_DIR / "uploads"
DOWNLOADS_DIR = FRONTEND_DIR / "downloads"
JOBS_DIR = FRONTEND_DIR / "jobs"

V1_MAIN_PATH = V1_DIR / "main.py"

UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
JOBS_DIR.mkdir(parents=True, exist_ok=True)

logger = logging.getLogger(__name__)


# ── Constants ─────────────────────────────────────────────────────────────────

VALID_POLICY_YEARS = {"2024", "2025", "2026"}

DEFAULT_REQUESTED_TABLES: dict[str, bool] = {
    "carry_forward_losses": False,
    "rd_tax_incentive": False,
    "div7a": False,
    "fbt_entertainment": False,
    "depreciation": False,
    "superannuation": False,
    "gst_reconciliation": False,
    "related_party_loans": False,
    "psi": False,
}

# ── Basic helpers ─────────────────────────────────────────────────────────────

def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _as_list(value: Any) -> list[Any]:
    """
    Streamlit uploader can return:
    - None
    - one UploadedFile
    - list[UploadedFile]
    """
    if value is None:
        return []

    if isinstance(value, (list, tuple)):
        return list(value)

    return [value]


def _safe_filename(name: str, fallback: str = "uploaded.xlsx") -> str:
    name = name or fallback
    safe = "".join(
        c if c.isalnum() or c in " ._-()" else "_"
        for c in name
    ).strip()

    return safe or fallback


def _safe_client_name(client_name: str) -> str:
    safe = "".join(
        c if c.isalnum() or c in "-_" else "_"
        for c in (client_name or "")
    ).strip("_")

    return safe


def _save_uploaded_file(uploaded_file: BinaryIO, dest_dir: Path, index: int) -> Path:
    original_name = getattr(uploaded_file, "name", f"uploaded_{index}.xlsx")
    safe_name = _safe_filename(original_name, fallback=f"uploaded_{index}.xlsx")
    dest = dest_dir / f"{index:02d}_{safe_name}"

    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    with open(dest, "wb") as f:
        f.write(uploaded_file.read())

    logger.info("Saved frontend upload: %s", dest)
    return dest


def _copy_backend_output_to_downloads(
    backend_output: Path,
    client_name: str = "",
    history_owner_id: str = "local",
) -> tuple[Path, str]:
    safe_client = _safe_client_name(client_name)
    prefix = f"{safe_client}_" if safe_client else ""
    output_name = f"{prefix}workpaper_{_timestamp()}.xlsx"

    safe_owner = _safe_client_name(history_owner_id) or "local"
    owner_download_dir = DOWNLOADS_DIR / safe_owner
    owner_download_dir.mkdir(parents=True, exist_ok=True)
    frontend_output = owner_download_dir / output_name
    shutil.copy2(backend_output, frontend_output)

    logger.info("Copied backend output to frontend downloads: %s", frontend_output)
    return frontend_output, output_name


def _extract_warnings(stdout: str, stderr: str) -> list[str]:
    text = f"{stdout}\n{stderr}"
    warnings: list[str] = []

    for line in text.splitlines():
        upper = line.upper()

        if "WARNING" in upper or "[WARNING]" in upper:
            stripped = line.strip()
            if stripped:
                warnings.append(stripped)

    return warnings[:30]


def _detect_reports_from_log(stdout: str, stderr: str) -> dict[str, bool]:
    text = f"{stdout}\n{stderr}".lower()

    return {
        "P&L": (
            "profit_and_loss" in text
            or "profit and loss" in text
            or "p&l" in text
        ),
        "Balance Sheet": (
            "balance_sheet" in text
            or "balance sheet" in text
        ),
    }


# ── Job option helpers ────────────────────────────────────────────────────────

def _normalise_requested_tables(
    requested_tables: dict[str, bool] | None,
) -> dict[str, bool]:
    tables = DEFAULT_REQUESTED_TABLES.copy()

    if not requested_tables:
        return tables

    for key, value in requested_tables.items():
        if key in tables:
            tables[key] = bool(value)

    return tables


def _normalise_policy_year(ato_policy_year: str = "2026") -> str:
    year = str(ato_policy_year or "2026").strip()

    if year not in VALID_POLICY_YEARS:
        return "2026"

    return year


def build_base_rate_entity_assessment(
    income_year: str,
    *,
    aggregated_turnover: float | int | str,
    total_assessable_income: float | int | str,
    base_rate_entity_passive_income: float | int | str,
    reviewer_confirmed: bool = False,
) -> dict[str, Any]:
    """Return a JSON-safe assessment for the frontend and job audit record."""

    assessment = assess_base_rate_entity(
        _normalise_policy_year(income_year),
        aggregated_turnover=aggregated_turnover,
        total_assessable_income=total_assessable_income,
        base_rate_entity_passive_income=base_rate_entity_passive_income,
    )
    return {
        "income_year": assessment.income_year,
        "aggregated_turnover": str(aggregated_turnover),
        "total_assessable_income": str(total_assessable_income),
        "base_rate_entity_passive_income": str(base_rate_entity_passive_income),
        "passive_income_ratio": str(assessment.passive_income_ratio),
        "turnover_threshold": str(assessment.turnover_threshold),
        "passive_income_ratio_limit": str(assessment.passive_income_ratio_limit),
        "turnover_below_threshold": assessment.turnover_below_threshold,
        "passive_income_ratio_within_limit": (
            assessment.passive_income_ratio_within_limit
        ),
        "eligible_on_supplied_figures": assessment.eligible_on_supplied_figures,
        "reviewer_confirmed": reviewer_confirmed is True,
    }


def _base_rate_assessment_is_confirmed(
    assessment: dict[str, Any] | None,
    income_year: str,
) -> bool:
    if not assessment or assessment.get("reviewer_confirmed") is not True:
        return False
    try:
        if to_decimal(
            assessment.get("total_assessable_income"),
            "total_assessable_income",
        ) <= 0:
            return False
        verified = build_base_rate_entity_assessment(
            income_year,
            aggregated_turnover=assessment.get("aggregated_turnover"),
            total_assessable_income=assessment.get("total_assessable_income"),
            base_rate_entity_passive_income=assessment.get(
                "base_rate_entity_passive_income"
            ),
            reviewer_confirmed=True,
        )
    except (CalculatorError, TypeError, ValueError):
        return False
    return verified["eligible_on_supplied_figures"] is True


def _build_job_options(
    ato_policy_year: str = "2026",
    requested_tables: dict[str, bool] | None = None,
    reviewer_notes: str = "",
    company_profile: str = "",
    document_description: str = "",
    client_name: str = "",
    company_tax_rate_category: str = "review_required",
    base_rate_entity_assessment: dict[str, Any] | None = None,
    retain_job_files: bool = False,
) -> dict[str, Any]:
    year = _normalise_policy_year(ato_policy_year)
    rate_category = company_tax_rate_category
    if rate_category == "base_rate_entity" and not _base_rate_assessment_is_confirmed(
        base_rate_entity_assessment,
        year,
    ):
        rate_category = "review_required"

    return {
        "ato_policy_year": year,
        "itr_policy_year": year,
        "requested_tables": _normalise_requested_tables(requested_tables),
        "reviewer_notes": reviewer_notes or "",
        "company_profile": company_profile or "",
        "document_description": document_description or "",
        "client_name": client_name or "",
        "company_tax_rate_category": (
            rate_category
            if rate_category in {"base_rate_entity", "general"}
            else "review_required"
        ),
        "base_rate_entity_assessment": base_rate_entity_assessment or {},
        "retain_job_files": bool(retain_job_files),
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "source": "frontend/job_runner.py",
    }


def _write_job_config(job_options: dict[str, Any], path: Path) -> Path:
    """
    Writes frontend-selected settings for v1/main.py, itr_rules.py, ato_policy.py,
    workpaper_builder.py, write_workbook.py, etc.

    Backend reads the per-job config path supplied below.

    Backend subprocess also receives:
        ATO_POLICY_YEAR
        ITR_POLICY_YEAR
        TAX_JOB_CONFIG_PATH
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(job_options, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    logger.info("Wrote backend job config: %s", path)
    return path


# ── Backend subprocess ────────────────────────────────────────────────────────

def _run_v1_main(
    job_options: dict[str, Any],
    *,
    job_root: Path,
    input_dir: Path,
    output_path: Path,
    log_dir: Path,
    job_config_path: Path,
) -> subprocess.CompletedProcess:
    if not V1_MAIN_PATH.exists():
        raise FileNotFoundError(f"Cannot find backend main.py at: {V1_MAIN_PATH}")

    env = os.environ.copy()

    env["ATO_POLICY_YEAR"] = str(job_options.get("ato_policy_year", "2026"))
    env["ITR_POLICY_YEAR"] = str(job_options.get("itr_policy_year", "2026"))
    env["SELECTED_INCOME_YEAR"] = str(job_options.get("ato_policy_year", "2026"))
    env["COMPANY_TAX_RATE_CATEGORY"] = str(
        job_options.get("company_tax_rate_category", "review_required")
    )
    env["TAX_JOB_CONFIG_PATH"] = str(job_config_path)
    env["TAX_JOB_WORK_DIR"] = str(job_root)
    env["TAX_DATA_DIR"] = str(input_dir)
    env["TAX_OUTPUT_DIR"] = str(output_path.parent)
    env["TAX_OUTPUT_PATH"] = str(output_path)
    env["TAX_LOG_DIR"] = str(log_dir)

    return subprocess.run(
        [sys.executable, "-m", "v1.main"],
        cwd=str(ROOT_DIR),
        text=True,
        capture_output=True,
        check=False,
        env=env,
    )


# ── Gemini face-check helpers ─────────────────────────────────────────────────

def _summarise_workbook_for_ai(path: Path, max_rows_per_sheet: int = 40) -> str:
    """
    Creates a compact text summary of the generated workbook.

    This intentionally does not send the whole workbook file to Gemini.
    It sends:
    - workbook name
    - sheet names
    - dimensions
    - first non-empty rows per sheet
    """
    if load_workbook is None:
        return "Workbook summary unavailable because openpyxl could not be imported."

    if not path.exists():
        return f"Workbook summary unavailable because file does not exist: {path}"

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return f"Workbook summary unavailable because openpyxl failed to read it: {exc}"

    parts: list[str] = []

    parts.append(f"Workbook: {path.name}")
    parts.append(f"Sheets: {', '.join(wb.sheetnames)}")

    for ws in wb.worksheets:
        parts.append("")
        parts.append(f"=== Sheet: {ws.title} ===")
        parts.append(f"Max rows: {ws.max_row}, max columns: {ws.max_column}")

        rows_added = 0

        for row in ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row, max_rows_per_sheet),
            values_only=True,
        ):
            values = ["" if cell is None else str(cell) for cell in row]
            joined = " | ".join(values).strip()

            if joined:
                parts.append(joined[:1200])
                rows_added += 1

            if rows_added >= max_rows_per_sheet:
                break

    try:
        wb.close()
    except Exception:
        pass

    return "\n".join(parts)[:60000]


def _run_gemini_face_check(
    job_options: dict[str, Any],
    backend_output: Path,
    stdout: str = "",
    stderr: str = "",
    api_key: str = "",
    model: str = "gemini-2.5-flash",
) -> dict[str, str]:
    """
    Optional AI review pass.

    Requires:
        pip install -U google-genai

    The API key and model are supplied explicitly by the current request.

    This is a face-check only. It should not replace accountant review.
    """
    api_key = str(api_key or "").strip()
    if not api_key:
        return {
            "status": "skipped",
            "summary": (
                "Gemini face-check skipped because no API key was supplied for "
                "this request."
            ),
        }

    try:
        from google import genai
    except Exception as exc:
        return {
            "status": "skipped",
            "summary": (
                "Gemini face-check skipped because google-genai is not installed. "
                "Install with: pip install -U google-genai\n\n"
                f"Import error: {exc}"
            ),
        }

    workbook_summary = _summarise_workbook_for_ai(backend_output)

    prompt = f"""
You are reviewing an Australian company tax workpaper generated from Xero reports.

Scope:
- This is an accountant-review face-check only.
- Do not provide final tax advice.
- Do not invent missing facts.
- Only flag issues visible from the user inputs, backend logs, and workbook summary.

Your job:
- Identify obvious issues on the face of the workpaper.
- Check whether selected optional schedules appear relevant.
- Check whether income and expense labels look unusual.
- Check whether these matters need review:
  - R&D tax incentive
  - carry-forward losses
  - Division 7A / shareholder loans
  - superannuation timing
  - depreciation / capital allowances
  - GST / BAS reconciliation
  - PSI
  - related-party loans
  - FBT / entertainment

User-selected job options:
{json.dumps(job_options, indent=2, ensure_ascii=False)}

Backend stdout tail:
{stdout[-8000:]}

Backend stderr tail:
{stderr[-8000:]}

Generated workbook summary:
{workbook_summary}

Return format:
1. Overall face-check result
2. Possible issues
3. Missing information / schedules to request
4. High-priority review points
5. Low-priority observations
"""

    try:
        client = genai.Client(api_key=api_key)

        response = client.models.generate_content(
            model=model or "gemini-2.5-flash",
            contents=prompt,
        )

        return {
            "status": "success",
            "summary": response.text or "Gemini returned an empty response.",
        }

    except Exception as exc:
        logger.exception("Gemini face-check failed")
        return {
            "status": "error",
            "summary": f"Gemini face-check failed: {type(exc).__name__}: {exc}",
        }


# ── Error helper ──────────────────────────────────────────────────────────────

def _build_error_result(
    base_result: dict[str, Any],
    message: str,
    stdout: str = "",
    stderr: str = "",
) -> dict[str, Any]:
    full_log = f"{stdout}\n{stderr}".strip()

    if full_log:
        base_result["error_message"] = f"{message}\n\n{full_log}"
    else:
        base_result["error_message"] = message

    base_result["warnings"] = _extract_warnings(stdout, stderr)
    base_result["detected"] = _detect_reports_from_log(stdout, stderr)
    base_result["backend_log"] = full_log

    return base_result


# ── Public function called by app.py ──────────────────────────────────────────

def run_workpaper_job(
    extra_files: BinaryIO | list[BinaryIO] | None = None,
    pl_file: BinaryIO | list[BinaryIO] | None = None,
    bs_file: BinaryIO | list[BinaryIO] | None = None,
    combined_file: BinaryIO | list[BinaryIO] | None = None,
    company_profile: str = "",
    document_description: str = "",
    client_name: str = "",
    ato_policy_year: str = "2026",
    requested_tables: dict[str, bool] | None = None,
    reviewer_notes: str = "",
    run_ai_face_check: bool = False,
    company_tax_rate_category: str = "review_required",
    base_rate_entity_assessment: dict[str, Any] | None = None,
    history_owner_id: str = "local",
    ai_provider: str = "None",
    ai_model: str = "",
    ai_api_key: str = "",
    retain_job_files: bool = False,
) -> dict[str, Any]:
    """
    Run the working v1 backend from Streamlit uploads.

    Recommended app.py call:
        run_workpaper_job(
            extra_files=uploaded_files,
            company_profile=company_profile,
            document_description=document_description,
            client_name=client_name,
            ato_policy_year=ato_policy_year,
            requested_tables=requested_tables,
            reviewer_notes=reviewer_notes,
            run_ai_face_check=run_ai_face_check,
        )
    """
    result: dict[str, Any] = {
        "status": "error",
        "output_path": None,
        "output_name": "",
        "detected": {},
        "review_count": None,
        "warnings": [],
        "uploaded_files": [],
        "frontend_upload_paths": [],
        "backend_data_paths": [],
        "backend_output_path": None,
        "backend_command": "",
        "backend_log": "",
        "company_profile": company_profile or "",
        "document_description": document_description or "",
        "client_name": client_name or "",
        "job_options": {},
        "job_config_path": "",
        "ai_face_check": None,
        "error_message": None,
        "job_id": "",
        "job_cleaned_up": False,
    }

    job_root: Path | None = None

    try:
        # 0. Build frontend-selected job options.
        job_options = _build_job_options(
            ato_policy_year=ato_policy_year,
            requested_tables=requested_tables,
            reviewer_notes=reviewer_notes,
            company_profile=company_profile,
            document_description=document_description,
            client_name=client_name,
            company_tax_rate_category=company_tax_rate_category,
            base_rate_entity_assessment=base_rate_entity_assessment,
            retain_job_files=retain_job_files,
        )

        result["job_options"] = job_options

        uploads: list[Any] = []
        uploads.extend(_as_list(extra_files))
        uploads.extend(_as_list(combined_file))
        uploads.extend(_as_list(pl_file))
        uploads.extend(_as_list(bs_file))

        if not uploads:
            result["error_message"] = "Please upload at least one Excel workbook."
            return result

        # 1. Save frontend uploads into timestamped folder.
        job_id = f"{_timestamp()}_{uuid.uuid4().hex}"
        result["job_id"] = job_id
        job_root = JOBS_DIR / job_id
        job_upload_dir = job_root / "inputs"
        job_output_dir = job_root / "output"
        job_log_dir = job_root / "logs"
        job_config_path = job_root / "job_config.json"
        backend_output = job_output_dir / "tax_workpaper.xlsx"
        job_upload_dir.mkdir(parents=True, exist_ok=True)
        job_output_dir.mkdir(parents=True, exist_ok=True)
        job_log_dir.mkdir(parents=True, exist_ok=True)

        saved_inputs = [
            _save_uploaded_file(uploaded_file, job_upload_dir, index)
            for index, uploaded_file in enumerate(uploads, start=1)
        ]

        result["uploaded_files"] = [path.name for path in saved_inputs]
        result["frontend_upload_paths"] = [str(path) for path in saved_inputs]

        # Inputs already live in this job's isolated backend data directory.
        result["backend_data_paths"] = [str(path) for path in saved_inputs]

        # 3b. Write frontend-selected tax settings for backend.
        job_config_path = _write_job_config(job_options, job_config_path)
        result["job_config_path"] = str(job_config_path)

        # 4. Run the backend package with this job's isolated paths.
        result["backend_command"] = f"{sys.executable} -m v1.main"

        completed = _run_v1_main(
            job_options,
            job_root=job_root,
            input_dir=job_upload_dir,
            output_path=backend_output,
            log_dir=job_log_dir,
            job_config_path=job_config_path,
        )

        stdout = completed.stdout or ""
        stderr = completed.stderr or ""
        full_log = f"{stdout}\n{stderr}".strip()

        result["backend_log"] = full_log
        result["warnings"] = _extract_warnings(stdout, stderr)
        result["detected"] = _detect_reports_from_log(stdout, stderr)

        if completed.returncode != 0:
            return _build_error_result(
                result,
                f"Backend exited with code {completed.returncode}.",
                stdout,
                stderr,
            )

        # 5. Use the output owned by this job only.
        if not backend_output.exists():
            return _build_error_result(
                result,
                "Backend completed, but this job's Excel output was not found.",
                stdout,
                stderr,
            )

        result["backend_output_path"] = str(backend_output)

        # 5b. Optional Gemini face-check before copying final download.
        if run_ai_face_check:
            if ai_provider != "Gemini":
                result["ai_face_check"] = {
                    "status": "skipped",
                    "summary": "Only Gemini face-check is currently supported.",
                }
            else:
                result["ai_face_check"] = _run_gemini_face_check(
                    job_options=job_options,
                    backend_output=backend_output,
                    stdout=stdout,
                    stderr=stderr,
                    api_key=ai_api_key,
                    model=ai_model,
                )
        else:
            result["ai_face_check"] = {
                "status": "skipped",
                "summary": "Gemini face-check was not selected.",
            }

        # 6. Copy backend output into frontend/downloads.
        frontend_output, output_name = _copy_backend_output_to_downloads(
            backend_output,
            client_name=client_name,
            history_owner_id=history_owner_id,
        )

        result.update(
            {
                "status": "success",
                "output_path": frontend_output,
                "output_name": output_name,
                "error_message": None,
            }
        )

        return result

    except Exception as exc:
        logger.exception("job_runner error")

        result["error_message"] = (
            f"{type(exc).__name__}: {exc}\n\n{traceback.format_exc()}"
        )

        return result

    finally:
        if job_root is not None and job_root.exists() and not retain_job_files:
            try:
                shutil.rmtree(job_root)
                result["job_cleaned_up"] = True
            except Exception:
                logger.exception("Could not clean isolated job directory: %s", job_root)
