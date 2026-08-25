"""Controlled, auditable edits to generated workpaper review fields.

This is intentionally not a general Excel editor. Source amounts, formulas and
rule evidence remain read-only. A reviewer may amend visible ITR references,
confidence and review notes; every accepted change is written to a new workbook
and an adjacent audit record.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EDITABLE_FIELDS = ("ITR Ref", "Confidence", "Review note")
READ_ONLY_FIELDS = ("Excel row", "Account", "Tab 3 decision", "Mapping reason")
VALID_CONFIDENCE = {"", "low", "medium", "high"}
_ITR_REF_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9 .()/-]{0,79}$")


class RevisionEditorError(ValueError):
    """Raised when a requested web revision is not safe to export."""


@dataclass(frozen=True)
class ReviewSheet:
    sheet_name: str
    header_row: int
    columns: dict[str, int]
    account_column: int | None


@dataclass(frozen=True)
class RevisionExport:
    workbook_path: Path
    audit_path: Path
    changes: tuple[dict[str, Any], ...]


def list_review_sheets(workbook_path: Path) -> tuple[ReviewSheet, ...]:
    """Find report sheets that contain the generated side-label review fields."""

    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        found: list[ReviewSheet] = []
        for worksheet in workbook.worksheets:
            sheet = _find_review_sheet(worksheet)
            if sheet is not None:
                found.append(sheet)
        return tuple(found)
    finally:
        workbook.close()


def load_review_rows(workbook_path: Path, sheet: ReviewSheet) -> list[dict[str, Any]]:
    """Return display rows with stable Excel row identities for the web grid."""

    workbook = load_workbook(Path(workbook_path), read_only=True, data_only=False)
    try:
        worksheet = workbook[sheet.sheet_name]
        rows: list[dict[str, Any]] = []
        for row_number in range(sheet.header_row + 1, worksheet.max_row + 1):
            account = _cell_text(worksheet.cell(row_number, sheet.account_column)) if sheet.account_column else ""
            values = {
                field: _cell_text(worksheet.cell(row_number, column))
                for field, column in sheet.columns.items()
            }
            # A source-account description distinguishes source rows from the
            # blank tail beneath a copied Xero report. A populated review field
            # is also enough to preserve a relevant structural/review row.
            if not account and not any(values.values()):
                continue
            if not account:
                account = "(structural or review row)"
            rows.append(
                {
                    "Excel row": row_number,
                    "Account": account,
                    "ITR Ref": values.get("ITR Ref", ""),
                    "Confidence": values.get("Confidence", ""),
                    "Tab 3 decision": values.get("Tab 3 decision", ""),
                    "Mapping reason": values.get("Mapping reason", ""),
                    "Review note": values.get("Review note", ""),
                }
            )
        return rows
    finally:
        workbook.close()


def collect_revision_changes(
    original_rows: list[dict[str, Any]],
    edited_rows: list[dict[str, Any]],
) -> tuple[dict[str, Any], ...]:
    """Validate and return only reviewer-selected editable-field changes."""

    original_by_row = {int(row["Excel row"]): row for row in original_rows}
    changes: list[dict[str, Any]] = []
    for edited in edited_rows:
        try:
            excel_row = int(edited["Excel row"])
        except (KeyError, TypeError, ValueError) as exc:
            raise RevisionEditorError("Every edited row must retain its Excel row identity.") from exc

        original = original_by_row.get(excel_row)
        if original is None:
            raise RevisionEditorError(f"Excel row {excel_row} is not part of this review grid.")

        for field in EDITABLE_FIELDS:
            old_value = _normalise_cell_value(original.get(field, ""))
            new_value = _normalise_cell_value(edited.get(field, ""))
            if old_value == new_value:
                continue
            _validate_change(field, new_value)
            changes.append(
                {
                    "sheet_row": excel_row,
                    "account": _normalise_cell_value(original.get("Account", "")),
                    "field": field,
                    "old_value": old_value,
                    "new_value": new_value,
                }
            )
    return tuple(changes)


def export_revision(
    *,
    source_workbook: Path,
    sheet: ReviewSheet,
    changes: tuple[dict[str, Any], ...],
    reviewer: str,
    revision_note: str,
) -> RevisionExport:
    """Copy, amend and audit a reviewed workpaper without touching its source."""

    source_workbook = Path(source_workbook)
    if not source_workbook.is_file():
        raise RevisionEditorError("The selected source workpaper no longer exists.")
    if not changes:
        raise RevisionEditorError("Change at least one editable review field before exporting.")

    reviewer = _normalise_cell_value(reviewer)
    revision_note = _normalise_cell_value(revision_note)
    if not reviewer:
        raise RevisionEditorError("Reviewer name or initials are required for an exported revision.")
    if not revision_note:
        raise RevisionEditorError("Add a short reason for this revision before exporting.")
    _reject_formula_like_text(reviewer, "Reviewer")
    _reject_formula_like_text(revision_note, "Revision note")

    output_path = _new_revision_path(source_workbook)
    shutil.copy2(source_workbook, output_path)
    workbook = load_workbook(output_path, data_only=False)
    try:
        worksheet = workbook[sheet.sheet_name]
        for change in changes:
            field = str(change["field"])
            column = sheet.columns.get(field)
            if column is None:
                raise RevisionEditorError(f"{field!r} is not editable in {sheet.sheet_name!r}.")
            cell = worksheet.cell(int(change["sheet_row"]), column)
            current_value = _normalise_cell_value(cell.value)
            if current_value != change["old_value"]:
                raise RevisionEditorError(
                    f"{sheet.sheet_name} row {change['sheet_row']} changed since it was opened. "
                    "Refresh the editor before exporting."
                )
            cell.value = change["new_value"]

        # openpyxl does not evaluate formulas. These flags ensure that Excel or
        # another compatible spreadsheet application recalculates formula-led
        # ITR summaries when the reviewer opens the revision.
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        workbook.save(output_path)
    except Exception:
        output_path.unlink(missing_ok=True)
        raise
    finally:
        workbook.close()

    audit_path = revision_audit_path(output_path)
    audit = {
        "version": 1,
        "kind": "manual_workpaper_revision",
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "source_workbook": source_workbook.name,
        "source_sha256": _sha256_file(source_workbook),
        "revision_workbook": output_path.name,
        "sheet_name": sheet.sheet_name,
        "reviewer": reviewer,
        "revision_note": revision_note,
        "changes": list(changes),
        "calculation_note": (
            "Formula cells are preserved and requested to recalculate when the revision is opened. "
            "This manual revision is not a backend rerun or an automatic tax conclusion."
        ),
    }
    _write_json_atomically(audit_path, audit)
    return RevisionExport(workbook_path=output_path, audit_path=audit_path, changes=changes)


def revision_audit_path(workbook_path: Path) -> Path:
    return Path(workbook_path).with_suffix(".revision_audit.json")


def _find_review_sheet(worksheet) -> ReviewSheet | None:
    max_scan_row = min(worksheet.max_row, 40)
    for row_number in range(1, max_scan_row + 1):
        headers = {
            _normalise_header(worksheet.cell(row_number, col_number).value): col_number
            for col_number in range(1, worksheet.max_column + 1)
            if _normalise_header(worksheet.cell(row_number, col_number).value)
        }
        if "itr ref" not in headers or "review note" not in headers:
            continue
        account_column = next(
            (headers[name] for name in ("account", "account name", "description") if name in headers),
            None,
        )
        return ReviewSheet(
            sheet_name=worksheet.title,
            header_row=row_number,
            columns={
                field: headers[field.casefold()]
                for field in ("ITR Ref", "Confidence", "Tab 3 decision", "Mapping reason", "Review note")
                if field.casefold() in headers
            },
            account_column=account_column,
        )
    return None


def _validate_change(field: str, value: str) -> None:
    if field == "ITR Ref":
        if not value:
            raise RevisionEditorError("ITR Ref cannot be blank. Use 'Review' when no filing reference is proposed.")
        if not _ITR_REF_RE.fullmatch(value):
            raise RevisionEditorError("ITR Ref contains unsupported characters or is too long.")
    elif field == "Confidence":
        if value.casefold() not in VALID_CONFIDENCE:
            raise RevisionEditorError("Confidence must be blank, low, medium or high.")
    elif field == "Review note":
        if len(value) > 2_000:
            raise RevisionEditorError("Review note must be 2,000 characters or fewer.")
        _reject_formula_like_text(value, "Review note")
    else:
        raise RevisionEditorError(f"{field!r} is not an editable review field.")


def _new_revision_path(source_workbook: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    candidate = source_workbook.with_name(f"{source_workbook.stem}_revision_{timestamp}.xlsx")
    while candidate.exists():
        candidate = source_workbook.with_name(
            f"{source_workbook.stem}_revision_{timestamp}_{uuid.uuid4().hex[:6]}.xlsx"
        )
    return candidate


def _normalise_header(value: object) -> str:
    return " ".join(str(value or "").strip().casefold().split())


def _normalise_cell_value(value: object) -> str:
    return str(value or "").strip()


def _cell_text(cell) -> str:
    return _normalise_cell_value(cell.value)


def _reject_formula_like_text(value: str, label: str) -> None:
    if value.startswith(("=", "+", "-", "@")):
        raise RevisionEditorError(f"{label} cannot start with an Excel formula character.")


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _write_json_atomically(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    try:
        temporary.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)
