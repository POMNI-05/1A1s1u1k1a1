"""Full-workbook spreadsheet canvas and manual-revision export helpers."""

from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

import streamlit as st
from openpyxl import load_workbook


class WorkbookCanvasError(ValueError):
    """Raised when a free workbook edit cannot be applied safely."""


_CANVAS_HTML = """
<div id="book"><div id="formula-bar"><span id="cell-address">A1</span><span class="fx">fx</span><input id="formula-input" spellcheck="false" /></div><div id="grid-wrap"><table id="grid"></table></div><div id="sheet-tabs"></div></div>
"""
_CANVAS_CSS = """
:host { display:block; height:100%; }
#book { height: calc(100vh - 185px); min-height: 600px; display:flex; flex-direction:column; border:1px solid #b8c2cc; background:#fff; font-family:Arial,Helvetica,sans-serif; overflow:hidden; }
#formula-bar { height:34px; display:flex; align-items:center; gap:7px; padding:0 8px; border-bottom:1px solid #cfd8df; background:#f7f8fa; }
#cell-address { width:62px; text-align:center; font-size:12px; color:#334155; border-right:1px solid #cfd8df; }
.fx { font-family:Georgia,serif; font-style:italic; color:#64748b; font-weight:bold; }
#formula-input { flex:1; min-width:0; height:23px; border:1px solid #cbd5e1; padding:0 7px; font-size:13px; background:#fff; }
#grid-wrap { flex:1; overflow:auto; background:white; }
table { border-collapse:separate; border-spacing:0; width:max-content; min-width:100%; font-size:13px; color:#17202a; }
th,td { height:25px; min-width:96px; max-width:360px; padding:0 6px; border-right:1px solid #dfe5ea; border-bottom:1px solid #dfe5ea; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; box-sizing:border-box; background:white; }
thead th { position:sticky; top:0; z-index:3; height:26px; text-align:center; font-size:11px; font-weight:500; color:#475569; background:#f1f3f5; }
.row { position:sticky; left:0; z-index:2; width:46px; min-width:46px; max-width:46px; padding:0 5px; text-align:right; font-size:11px; color:#64748b; background:#f6f7f9; }
thead .row { z-index:5; }
td { cursor:cell; } td:focus, td.selected { outline:2px solid #217346; outline-offset:-2px; z-index:1; overflow:visible; white-space:normal; background:#fff; }
#sheet-tabs { height:33px; display:flex; align-items:end; gap:2px; padding:0 8px; border-top:1px solid #cfd8df; background:#f7f8fa; overflow-x:auto; }
.tab { cursor:pointer; border:1px solid transparent; border-bottom:0; padding:8px 14px 7px; font-size:12px; color:#334155; white-space:nowrap; }
.tab.active { background:white; border-color:#cfd8df; border-top:3px solid #217346; padding-top:5px; font-weight:600; }
"""
_CANVAS_JS = """
export default function({ parentElement, data, setStateValue }) {
 const sheets = data?.sheets || []; const book = parentElement.querySelector('#book'); const grid = parentElement.querySelector('#grid'); const tabs = parentElement.querySelector('#sheet-tabs'); const address = parentElement.querySelector('#cell-address'); const formula = parentElement.querySelector('#formula-input');
 const edits = new Map(Object.entries(data?.edits || {})); let active = data?.active_sheet || sheets[0]?.name || ''; let selected = null;
 const keyFor = (sheet,r,c) => `${sheet}::${r}::${c}`;
 const colName = (n) => { let s=''; n++; while(n){let x=(n-1)%26;s=String.fromCharCode(65+x)+s;n=Math.floor((n-1)/26)} return s; };
 const cellValue = (sheet,r,c,base) => edits.has(keyFor(sheet,r,c)) ? edits.get(keyFor(sheet,r,c)) : (base ?? '');
 const save = () => setStateValue('edits', Object.fromEntries(edits));
 const redraw = () => { const sheet = sheets.find(x => x.name === active) || sheets[0]; if (!sheet) return; active = sheet.name; grid.replaceChildren(); tabs.replaceChildren();
   const head=document.createElement('thead'), hr=document.createElement('tr'), corner=document.createElement('th'); corner.className='row'; hr.appendChild(corner); for(let c=0;c<sheet.max_col;c++){const th=document.createElement('th');th.textContent=colName(c);hr.appendChild(th)} head.appendChild(hr);grid.appendChild(head);
   const body=document.createElement('tbody'); for(let r=0;r<sheet.max_row;r++){const tr=document.createElement('tr'), num=document.createElement('td');num.className='row';num.textContent=r+1;tr.appendChild(num);for(let c=0;c<sheet.max_col;c++){const td=document.createElement('td');td.contentEditable='true';td.spellcheck=false;td.dataset.r=r;td.dataset.c=c;td.textContent=cellValue(active,r,c,sheet.rows[r]?.[c]);td.title=td.textContent;
     const select=()=>{if(selected)selected.classList.remove('selected');selected=td;td.classList.add('selected');address.textContent=`${colName(c)}${r+1}`;formula.value=td.textContent;}; td.onfocus=select;td.onclick=select;
     td.onblur=()=>{const next=td.textContent.trimEnd(),base=String(sheet.rows[r]?.[c]??'');const k=keyFor(active,r,c);if(next===base)edits.delete(k);else edits.set(k,next);td.title=next;save();};
     td.onkeydown=(e)=>{if(e.key==='Tab'){e.preventDefault();td.blur();const next=tr.children[c+2]||body.rows[r+1]?.children[1];if(next)next.focus()}else if(e.key==='Enter'){e.preventDefault();td.blur();const next=body.rows[r+1]?.children[c+1];if(next)next.focus()}};
     td.onpaste=(e)=>{e.preventDefault();const matrix=(e.clipboardData.getData('text/plain')||'').replace(/\\r/g,'').split('\\n').map(x=>x.split('\\t'));matrix.forEach((line,dr)=>line.forEach((value,dc)=>{if(r+dr<sheet.max_row&&c+dc<sheet.max_col){const k=keyFor(active,r+dr,c+dc),base=String(sheet.rows[r+dr]?.[c+dc]??'');if(value===base)edits.delete(k);else edits.set(k,value)} }));save();redraw();};tr.appendChild(td)}body.appendChild(tr)}grid.appendChild(body);
   sheets.forEach(item=>{const tab=document.createElement('button');tab.className='tab'+(item.name===active?' active':'');tab.textContent=item.name;tab.onclick=()=>{active=item.name;setStateValue('active_sheet',active);redraw()};tabs.appendChild(tab)});
 };
 formula.onchange=()=>{if(!selected)return;selected.textContent=formula.value;selected.blur();}; redraw();
}
"""
_component = st.components.v2.component("full_workbook_spreadsheet", html=_CANVAS_HTML, css=_CANVAS_CSS, js=_CANVAS_JS)


def load_workbook_canvas(workbook_path: Path) -> list[dict[str, Any]]:
    """Load every visible sheet as display-safe values while preserving source files."""

    workbook = load_workbook(Path(workbook_path), data_only=False, read_only=True)
    try:
        sheets: list[dict[str, Any]] = []
        for worksheet in workbook.worksheets:
            if worksheet.sheet_state != "visible":
                continue
            rows = [
                ["" if value is None else str(value) for value in row]
                for row in worksheet.iter_rows(
                    min_row=1, max_row=max(worksheet.max_row, 1), min_col=1, max_col=max(worksheet.max_column, 1), values_only=True
                )
            ]
            sheets.append({"name": worksheet.title, "max_row": len(rows), "max_col": max(worksheet.max_column, 1), "rows": rows})
        return sheets
    finally:
        workbook.close()


def merge_workbook_edits(sheets: list[dict[str, Any]], edits: dict[str, Any] | None) -> dict[str, str]:
    """Keep only changes that target an existing visible workbook cell."""

    known = {sheet["name"]: sheet for sheet in sheets}
    valid: dict[str, str] = {}
    for key, value in (edits or {}).items():
        try:
            sheet_name, row_text, col_text = str(key).rsplit("::", 2)
            row, col = int(row_text), int(col_text)
        except (TypeError, ValueError):
            continue
        sheet = known.get(sheet_name)
        if sheet and 0 <= row < sheet["max_row"] and 0 <= col < sheet["max_col"]:
            original = str(sheet["rows"][row][col] or "")
            new = str(value or "")
            if new != original:
                valid[key] = new
    return valid


def render_workbook_canvas(sheets: list[dict[str, Any]], *, key: str) -> dict[str, str]:
    """Render a full-width free-edit spreadsheet and return its cell changes."""

    state = st.session_state.get(key, {})
    state = state if isinstance(state, dict) else {}
    prior_edits = merge_workbook_edits(sheets, state.get("edits"))
    active_sheet = str(state.get("active_sheet", sheets[0]["name"] if sheets else ""))
    result = _component(
        key=key,
        data={"sheets": sheets, "edits": prior_edits, "active_sheet": active_sheet},
        default={"edits": prior_edits, "active_sheet": active_sheet},
        on_edits_change=lambda: None,
        on_active_sheet_change=lambda: None,
        height="stretch",
    )
    return merge_workbook_edits(sheets, getattr(result, "edits", prior_edits))


def export_manual_workbook_revision(*, source_workbook: Path, sheets: list[dict[str, Any]], edits: dict[str, str]) -> tuple[Path, Path, int]:
    """Write unrestricted browser edits to a new workbook plus a change audit."""

    source_workbook = Path(source_workbook)
    if not source_workbook.is_file():
        raise WorkbookCanvasError("The selected workbook no longer exists.")
    changes = merge_workbook_edits(sheets, edits)
    if not changes:
        raise WorkbookCanvasError("No cell changes are waiting to be saved.")
    output = _revision_path(source_workbook)
    shutil.copy2(source_workbook, output)
    workbook = load_workbook(output, data_only=False)
    audit_changes = []
    try:
        for key, new_value in changes.items():
            sheet_name, row_text, col_text = key.rsplit("::", 2)
            row, col = int(row_text) + 1, int(col_text) + 1
            cell = workbook[sheet_name].cell(row, col)
            old_value = "" if cell.value is None else str(cell.value)
            cell.value = _coerce_cell_value(new_value)
            audit_changes.append({"sheet": sheet_name, "cell": cell.coordinate, "old_value": old_value, "new_value": new_value})
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        workbook.save(output)
    except Exception:
        output.unlink(missing_ok=True)
        raise
    finally:
        workbook.close()
    audit_path = output.with_suffix(".manual_edit_audit.json")
    audit = {"version": 1, "kind": "unrestricted_manual_web_workbook_edit", "created_at": datetime.now().isoformat(timespec="seconds"), "source_workbook": source_workbook.name, "source_sha256": _sha256(source_workbook), "revision_workbook": output.name, "change_count": len(audit_changes), "changes": audit_changes, "warning": "Manual browser edits may change formulas and tax outcomes. This file is not a deterministic backend rerun."}
    _write_json(audit_path, audit)
    return output, audit_path, len(audit_changes)


def _revision_path(source: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = source.with_name(f"{source.stem}_manual_edit_{stamp}.xlsx")
    while path.exists():
        path = source.with_name(f"{source.stem}_manual_edit_{stamp}_{uuid.uuid4().hex[:6]}.xlsx")
    return path


def _coerce_cell_value(value: str) -> Any:
    if value.startswith("="):
        return value
    if re.fullmatch(r"-?\d+", value):
        return int(value)
    if re.fullmatch(r"-?(?:\d+\.\d*|\d*\.\d+)", value):
        return float(value)
    return value


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _write_json(path: Path, data: dict[str, Any]) -> None:
    temp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    try:
        temp.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
        os.replace(temp, path)
    finally:
        temp.unlink(missing_ok=True)
