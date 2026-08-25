from __future__ import annotations

import io
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from frontend import job_runner
from ai_review import read_ai_review_audit


def _combined_workbook() -> io.BytesIO:
    workbook = Workbook()
    pl = workbook.active
    pl.title = "Profit and Loss"
    pl.append(["Profit and Loss"])
    pl.append([])
    pl.append(["Account", "30 Jun 2026"])
    pl.append(["Trading Income", None])
    pl.append(["Sales", 1000])
    pl.append(["Total Income", 1000])
    pl.append(["Operating Expenses", None])
    pl.append(["Entertainment", 100])
    pl.append(["Total Operating Expenses", 100])
    pl.append(["Net Profit", 900])

    bs = workbook.create_sheet("Balance Sheet")
    bs.append(["Balance Sheet"])
    bs.append([])
    bs.append(["Account", "30 Jun 2026"])
    bs.append(["Current Assets", None])
    bs.append(["Cash", 1000])
    bs.append(["Total Assets", 1000])
    bs.append(["Current Liabilities", None])
    bs.append(["Loan", 400])
    bs.append(["Total Liabilities", 400])
    bs.append(["Equity", None])
    bs.append(["Retained Earnings", 600])
    bs.append(["Total Equity", 600])
    bs.append(["Net Assets", 600])

    stream = io.BytesIO()
    stream.name = "combined.xlsx"
    workbook.save(stream)
    stream.seek(0)
    return stream


class EndToEndTests(unittest.TestCase):
    def test_isolated_job_generates_workbook_and_cleans_working_files(self):
        with TemporaryDirectory() as folder:
            root = Path(folder)
            with (
                patch.object(job_runner, "JOBS_DIR", root / "jobs"),
                patch.object(job_runner, "DOWNLOADS_DIR", root / "downloads"),
            ):
                result = job_runner.run_workpaper_job(
                    extra_files=[_combined_workbook()],
                    client_name="Synthetic",
                    ato_policy_year="2026",
                    company_tax_rate_category="general",
                    history_owner_id="test-session",
                    requested_tables={"fbt_entertainment": True},
                )

            self.assertEqual(result["status"], "success", result.get("error_message"))
            self.assertTrue(result["job_cleaned_up"])
            self.assertFalse((root / "jobs" / result["job_id"]).exists())
            self.assertEqual(result["workpaper_result"]["status"], "completed")
            self.assertGreater(result["workpaper_result"]["decision_trace_count"], 0)

            output = Path(result["output_path"])
            self.assertTrue(output.exists())
            self.assertEqual(output.parent.name, "test-session")
            audit = read_ai_review_audit(result["ai_review_audit_path"])
            self.assertEqual(audit["response"]["status"], "skipped")
            self.assertEqual(audit["provider"]["name"], "None")
            self.assertEqual(audit["accountant_disposition"]["status"], "pending")

            workbook = load_workbook(output, data_only=False)
            self.assertIn("Tax Reconciliation", workbook.sheetnames)
            self.assertNotIn("Tax Calculation", workbook.sheetnames)
            self.assertIn("Inputs & Overrides", workbook.sheetnames)
            recon_values = [
                cell.value for row in workbook["Tax Reconciliation"] for cell in row
            ]
            checks_values = [
                cell.value for row in workbook["Checks"] for cell in row
            ]
            input_values = [
                cell.value for row in workbook["Inputs & Overrides"] for cell in row
            ]
            self.assertIn(
                "Proposed Tax Adjustments - Not Posted Unless Approved",
                input_values,
            )
            self.assertIn("FBT / Entertainment Review", input_values)
            self.assertIn("ADD", recon_values)
            self.assertIn("Total ADD", recon_values)
            self.assertIn(
                "Preliminary taxable income/(loss) — Item 7T",
                recon_values,
            )
            # Net profit is 900. The proposed 100 entertainment add-back is
            # included in the Tab 3 preliminary calculation, then reviewed
            # before a final lodged Item 7T is prepared.
            recon_sheet = workbook["Tax Reconciliation"]
            base_row = next(
                row[0].row for row in recon_sheet.iter_rows()
                if row[0].value == "Accounting profit/(loss) — Item 6T"
            )
            add_total_row = next(
                row[0].row for row in recon_sheet.iter_rows()
                if row[0].value == "Total ADD"
            )
            result_row = next(
                row[0].row for row in recon_sheet.iter_rows()
                if row[0].value == "Preliminary taxable income/(loss) — Item 7T"
            )
            self.assertEqual(recon_sheet.cell(add_total_row, 2).value, "=SUM(B5:B5)")
            self.assertEqual(
                recon_sheet.cell(result_row, 2).value,
                f"=B{base_row}+B{add_total_row}-0",
            )
            self.assertNotIn(300, recon_values)
            self.assertNotIn("Completeness checks — not tax adjustments", recon_values)
            self.assertIn("Tax Reconciliation Review Checks", checks_values)

            add_heading = next(
                row[0] for row in recon_sheet.iter_rows()
                if row[0].value == "ADD"
            )
            self.assertTrue(str(add_heading.font.color.rgb).endswith("FF0000"))
            self.assertIsNone(add_heading.fill.fill_type)

            balance_sheet = workbook["Balance Sheet"]
            self.assertNotIn(
                "Tab 3 decision",
                [cell.value for row in balance_sheet.iter_rows() for cell in row],
            )


if __name__ == "__main__":
    unittest.main()
