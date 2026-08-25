from __future__ import annotations

import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook

from v1.cleaner import (
    AmountParseStatus,
    ReportInput,
    add_row_type,
    choose_current_amount_col,
    clean_report_input,
    extract_tax_depreciation_total,
    parse_amount,
)
from v1.workpaper_builder import (
    _build_bs_checks,
    _build_label_summary,
    _extract_net_profit_by_period,
)
from v1.labeller import (
    _apply_balance_sheet_section_conflict_review,
    _should_label_row,
)
from v1.write_workbook import _find_main_amount_col, _looks_like_generated_column


class MinimalSafetyRepairTests(unittest.TestCase):
    def test_requested_year_is_selected_instead_of_numeric_density(self):
        df = pd.DataFrame({"Account": ["Sales"], "2024": [10], "2025": [5]})
        with patch("v1.cleaner.SELECTED_INCOME_YEAR", "2025"):
            self.assertEqual(choose_current_amount_col(df, ["2024", "2025"]), "2025")

    def test_sales_and_other_income_with_amounts_are_accounts(self):
        df = pd.DataFrame(
            {
                "Account Label": ["Sales", "Other Income", "Income"],
                "2025": [100.0, 25.0, 0.0],
            }
        )
        result = add_row_type(df, "Account Label", ["2025"])
        self.assertEqual(result["Row Type"].tolist(), ["account", "account", "heading"])

    def test_gross_profit_with_an_amount_is_a_structural_total_not_an_account(self):
        df = pd.DataFrame(
            {
                "Account Label": ["Sales", "Cost of Sales", "Gross Profit"],
                "2025": [1000.0, 300.0, 700.0],
            }
        )

        result = add_row_type(df, "Account Label", ["2025"])

        self.assertEqual(result["Row Type"].tolist(), ["account", "account", "total"])

    def test_balance_sheet_summary_uses_itr_label_when_no_workpaper_label_exists(self):
        labelled = pd.DataFrame(
            {
                "Account": ["Total Current Liabilities", "Total Liabilities"],
                "2025": [120.0, 250.0],
                "Row Type": ["total", "total"],
                "ITR Ref": ["8G", "8H"],
                "ITR Label": ["All current liabilities", "Total liabilities"],
                "Treatment": ["financial_label_only", "financial_label_only"],
                "Confidence": ["high", "high"],
                "Review Note": ["", ""],
                "Source Row": [10, 12],
            }
        )

        summary = _build_label_summary(labelled, "BS")

        self.assertEqual(summary["ITR Ref"].tolist(), ["8G", "8H"])
        self.assertEqual(summary["Amount"].tolist(), [120.0, 250.0])

    def test_non_direct_balance_sheet_totals_do_not_receive_normal_labels(self):
        self.assertTrue(
            _should_label_row("total", "balance_sheet", "Total Current Liabilities")
        )
        self.assertFalse(_should_label_row("total", "balance_sheet", "Net Assets"))
        self.assertFalse(_should_label_row("total", "balance_sheet", "Total Equity"))

    def test_cash_account_under_liabilities_becomes_review_not_reclassification(self):
        mapping = _apply_balance_sheet_section_conflict_review(
            {
                "ITR Ref": "",
                "ITR Label": "Cash / bank support",
                "Treatment": "support_only",
                "Confidence": "high",
                "Label Reason": "Cash/bank supports current assets.",
            },
            account_name="Business Bank Account",
            report_section="Current Liabilities",
        )

        self.assertEqual(mapping["ITR Ref"], "Review")
        self.assertEqual(mapping["Treatment"], "review_only")
        self.assertEqual(mapping["Rule Source"], "structural_validation")

        unchanged = _apply_balance_sheet_section_conflict_review(
            {"Treatment": "support_only"},
            account_name="Bank Loan",
            report_section="Current Liabilities",
        )
        self.assertEqual(unchanged["Treatment"], "support_only")

    def test_explicit_excel_error_is_not_converted_to_zero(self):
        raw = pd.DataFrame(
            [
                ["Account", "2025"],
                ["Sales", "#REF!"],
                ["Other Income", 20],
            ]
        )
        report = ReportInput(
            report_type="profit_and_loss",
            source_path=Path("broken.xlsx"),
            sheet_name="P&L",
            raw_df=raw,
            detection_score=99,
            detection_reason="test",
        )
        with self.assertRaisesRegex(ValueError, "CELL-001"):
            clean_report_input(report)

    def test_unparseable_monetary_text_is_not_converted_to_zero(self):
        raw = pd.DataFrame(
            [
                ["Account", "2025"],
                ["Sales", "$12O0"],
                ["Other Income", 20],
            ]
        )
        report = ReportInput(
            report_type="profit_and_loss",
            source_path=Path("invalid-amount.xlsx"),
            sheet_name="P&L",
            raw_df=raw,
            detection_score=99,
            detection_reason="test",
        )
        with self.assertRaisesRegex(ValueError, "CELL-002"):
            clean_report_input(report)

    def test_xero_structural_column_is_not_parsed_as_a_monetary_column(self):
        raw = pd.DataFrame(
            [
                ["Column 1", "Account", "30 June 2026", "Column 4"],
                ["Assets", None, None, None],
                [None, "Bank", "100.00", "[FX]"],
                ["Total Assets", None, "100.00", None],
            ]
        )
        report = ReportInput(
            report_type="balance_sheet",
            source_path=Path("xero-balance-sheet.xlsx"),
            sheet_name="Balance Sheet",
            raw_df=raw,
            detection_score=99,
            detection_reason="test",
        )

        cleaned = clean_report_input(report)

        self.assertNotIn("Column 1 Parse Status", cleaned.columns)
        self.assertIn(100.0, cleaned["30 June 2026"].tolist())

    def test_amount_parse_result_preserves_blank_zero_and_valid_states(self):
        blank = parse_amount("")
        zero = parse_amount("0")
        valid = parse_amount("$12.50")

        self.assertEqual(blank.status, AmountParseStatus.BLANK)
        self.assertIsNone(blank.value)
        self.assertEqual(zero.status, AmountParseStatus.VALID)
        self.assertEqual(zero.value, 0.0)
        self.assertEqual(valid.status, AmountParseStatus.VALID)
        self.assertEqual(valid.value, 12.5)

    def test_cleaned_report_keeps_amount_parse_status(self):
        raw = pd.DataFrame(
            [
                ["Account", "2025"],
                ["Sales", None],
                ["Other Income", 0],
                ["Service Income", "$12.50"],
            ]
        )
        report = ReportInput(
            report_type="profit_and_loss",
            source_path=Path("parse-status.xlsx"),
            sheet_name="P&L",
            raw_df=raw,
            detection_score=99,
            detection_reason="test",
        )

        result = clean_report_input(report)

        self.assertEqual(
            result["2025 Parse Status"].tolist(),
            ["blank", "valid", "valid"],
        )

    def test_reuploaded_workpaper_ignores_generated_itr_summary_amounts(self):
        raw = pd.DataFrame(
            [
                [None, "30 Jun 2025", None, "ITR Label", None, "ITR Totals", None],
                ["Sales", 120.0, None, "Inc - 6C", None, "Inc - 6C", "=SUMIF(D:D,F2,B:B)"],
                ["Operating Expenses", None, None, None, None, None, None],
                ["Entertainment", 20.0, None, "Exp - 6S", None, "Exp - 6S", "=SUMIF(D:D,F4,B:B)"],
                ["Net Profit", 100.0, None, "6T", None, "PRE TAX PROFIT/(LOSS)", "=B5"],
            ]
        )
        report = ReportInput(
            report_type="profit_and_loss",
            source_path=Path("previous-workpaper.xlsx"),
            sheet_name="Profit and Loss",
            raw_df=raw,
            detection_score=99,
            detection_reason="test",
        )

        result = clean_report_input(report)

        self.assertIn("30 Jun 2025", result.columns)
        self.assertNotIn("Column 7", result.columns)
        self.assertEqual(result["30 Jun 2025"].iloc[[0, 2, 3]].tolist(), [120.0, 20.0, 100.0])
        self.assertTrue(result["30 Jun 2025"].iloc[1] != result["30 Jun 2025"].iloc[1])

    def test_known_side_by_side_tax_and_tb_layout_is_rejected(self):
        raw = pd.DataFrame(
            [
                ["Tax Return Disclosures", None, "Source Data from Client Trial Balance"],
                ["Account No", "Account Name", "GL Accounts"],
                ["C", "Gross Written Premiums", "Westpac Visa Account"],
            ]
        )
        report = ReportInput(
            report_type="profit_and_loss",
            source_path=Path("composite.xlsx"),
            sheet_name="ITR - INCOME",
            raw_df=raw,
            detection_score=99,
            detection_reason="test",
        )
        with self.assertRaisesRegex(ValueError, "STRUCT-003"):
            clean_report_input(report)

    def test_profit_priority_excludes_after_tax_and_falls_back_per_period(self):
        df = pd.DataFrame(
            {
                "Account Label": ["Profit Before Tax", "Net Profit After Tax", "Other Income"],
                "2025": [100.0, 80.0, 120.0],
                "2024": [0.0, 70.0, 110.0],
                "Row Type": ["total", "total", "account"],
                "Report Section": ["", "", "Income"],
            }
        )
        values, methods = _extract_net_profit_by_period(df)
        self.assertEqual(values["2025"], 100.0)
        self.assertEqual(values["2024"], 110.0)
        self.assertIn("Profit before tax", methods["2025"])
        self.assertIn("Fallback", methods["2024"])

    def test_depreciation_requires_explicit_total_and_deduction_column(self):
        raw = pd.DataFrame(
            [
                ["Tax Depreciation Schedule", None, None, None],
                ["Name", "Cost", "Depreciation", "Closing Value"],
                ["Asset A", 1000.0, 100.0, 900.0],
                ["Total", 1000.0, 100.0, 900.0],
            ]
        )
        report = ReportInput(
            report_type="tax_depreciation",
            source_path=Path("depreciation.xlsx"),
            sheet_name="Tax Depreciation",
            raw_df=raw,
            detection_score=99,
            detection_reason="test",
        )
        self.assertEqual(extract_tax_depreciation_total(report), 100.0)

    def test_missing_bs_total_is_not_tested_instead_of_zero(self):
        df = pd.DataFrame(
            {
                "Account Label": ["Total Assets", "Total Equity"],
                "2025": [100.0, 100.0],
                "Row Type": ["total", "total"],
            }
        )
        checks = _build_bs_checks(df)
        self.assertTrue(checks["Status"].str.startswith("NOT TESTED - BS-001").all())
        self.assertTrue(checks["2025"].isna().all())

    def test_account_column_is_not_removed_for_containing_total_rows(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Account"
        ws["A2"] = "Sales"
        ws["A3"] = "Total Income"
        ws["B1"] = "ITR Label"
        self.assertFalse(_looks_like_generated_column(ws, 1))
        self.assertTrue(_looks_like_generated_column(ws, 2))

    def test_writer_rejects_ambiguous_numeric_columns_without_requested_year(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Account", "Current", "Prior"])
        ws.append(["Sales", 100.0, 90.0])
        with self.assertRaisesRegex(ValueError, "PERIOD-001"):
            _find_main_amount_col(ws, 1, 2, 3)

    def test_writer_uses_period_in_account_header_not_year_in_report_title(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Profit and Loss", None, None])
        ws.append(["For the year ended 30 June 2025", None, None])
        ws.append(["Account", "2025", "2024"])
        ws.append(["Sales", 100.0, 90.0])

        with patch("v1.write_workbook.SELECTED_INCOME_YEAR", "2025"):
            self.assertEqual(_find_main_amount_col(ws, 1, 4, 3), 2)


if __name__ == "__main__":
    unittest.main()
