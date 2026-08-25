from __future__ import annotations

import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from types import SimpleNamespace
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook, load_workbook

from v1.cleaner import ReportInput
from v1.workpaper_builder import _build_carry_forward_losses_input, _manual_adjustment_rows
from v1.write_workbook import write_workbook


def _source_report(path: Path, title: str, periods: list[str], rows: list[tuple]) -> ReportInput:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append(["Account", *periods])
    for row in rows:
        sheet.append(list(row))
    workbook.save(path)
    return ReportInput(
        report_type="profit_and_loss" if title == "P&L" else "balance_sheet",
        source_path=path,
        sheet_name=title,
        raw_df=pd.DataFrame(),
        detection_score=99,
        detection_reason="test",
    )


def _workpaper(periods: list[str]):
    labelled_pl = pd.DataFrame([
        {"Source Row": 2, "Row Type": "account", "ITR Ref": "Inc - 6C", "Confidence": "high", "Label Reason": "Sales rule", "Review Note": "", "Treatment": "auto"}
    ])
    labelled_bs = pd.DataFrame([
        {"Source Row": 2, "Row Type": "account", "ITR Ref": "8A", "Confidence": "medium", "Label Reason": "Cash support", "Review Note": "Confirm presentation", "Treatment": "review_only"}
    ])
    recon_rows = [
        {"Line Type": "result", "Description": "Accounting Profit Before Tax", **{period: 100.0 for period in periods}, "ITR Ref": "6T", "Review note": ""},
        {"Line Type": "result", "Description": "Taxable Income", **{period: 100.0 for period in periods}, "ITR Ref": "7T", "Review note": ""},
    ]
    return SimpleNamespace(
        labelled_pl=labelled_pl,
        labelled_bs=labelled_bs,
        pl_label_summary=pd.DataFrame(),
        bs_label_summary=pd.DataFrame(),
        tax_reconciliation=pd.DataFrame(recon_rows),
        carry_forward_losses=pd.DataFrame(),
        rd_breakdown=pd.DataFrame(),
        bs_checks=pd.DataFrame(),
        review_items=pd.DataFrame(),
        proposed_adjustments=pd.DataFrame(),
        support_tables={},
    )


class OutputPresentationTests(unittest.TestCase):
    def test_reconciliation_renders_only_source_periods(self):
        for periods in (["2026"], ["2026", "2025"], ["2026", "2025", "2024"]):
            with self.subTest(periods=periods), TemporaryDirectory() as folder:
                root = Path(folder)
                pl_input = _source_report(root / "pl.xlsx", "P&L", list(periods), [("Sales", *([100.0] * len(periods)))])
                bs_input = _source_report(root / "bs.xlsx", "BS", list(periods), [("Cash", *([100.0] * len(periods)))])
                reports = SimpleNamespace(pl_input=pl_input, bs_input=bs_input, tax_depreciation_report=None, tax_depreciation_total=None, tax_depreciation_source=None)
                output = root / "output.xlsx"
                with patch("v1.write_workbook.OUTPUT_PATH", output):
                    write_workbook(reports, _workpaper(list(periods)))

                workbook = load_workbook(output, data_only=False)
                self.assertEqual(workbook.sheetnames, ["Profit and Loss", "Balance Sheet", "Tax Reconciliation", "Checks"])
                recon = workbook["Tax Reconciliation"]
                headers = [recon.cell(2, col).value for col in range(1, recon.max_column + 1)]
                rendered_years = [header for header in headers if str(header).startswith("Year Ended")]
                self.assertEqual(len(rendered_years), len(periods))
                values = [cell.value for row in recon.iter_rows() for cell in row]
                self.assertNotIn("Sales", values)
                self.assertNotIn("Cash", values)

    def test_fixed_assets_requires_validated_depreciation_evidence(self):
        with TemporaryDirectory() as folder:
            root = Path(folder)
            periods = ["2026"]
            pl_input = _source_report(root / "pl.xlsx", "P&L", periods, [("Sales", 100.0)])
            bs_input = _source_report(root / "bs.xlsx", "BS", periods, [("Cash", 100.0)])
            dep_input = _source_report(root / "dep.xlsx", "Depreciation", periods, [("Total tax depreciation", 25.0)])
            reports = SimpleNamespace(pl_input=pl_input, bs_input=bs_input, tax_depreciation_report=dep_input, tax_depreciation_total=25.0, tax_depreciation_source="dep.xlsx / Depreciation")
            output = root / "output.xlsx"
            with patch("v1.write_workbook.OUTPUT_PATH", output):
                write_workbook(reports, _workpaper(periods))

            workbook = load_workbook(output, data_only=False)
            self.assertIn("Fixed Assets", workbook.sheetnames)
            self.assertEqual(
                workbook.sheetnames,
                ["Profit and Loss", "Balance Sheet", "Fixed Assets", "Tax Reconciliation", "Checks"],
            )
            self.assertEqual(workbook["Fixed Assets"]["A2"].value, "Total tax depreciation")

    def test_raw_review_sheets_hide_tab3_decision_and_widen_notes(self):
        with TemporaryDirectory() as folder:
            root = Path(folder)
            periods = ["2026"]
            reports = SimpleNamespace(
                pl_input=_source_report(root / "pl.xlsx", "P&L", periods, [("Sales", 100.0)]),
                bs_input=_source_report(root / "bs.xlsx", "BS", periods, [("Cash", 100.0)]),
                tax_depreciation_report=None,
                tax_depreciation_total=None,
                tax_depreciation_source=None,
            )
            workpaper = _workpaper(periods)
            workpaper.labelled_pl.loc[0, "Review Note"] = "A long reviewer explanation."
            output = root / "output.xlsx"
            with patch("v1.write_workbook.OUTPUT_PATH", output):
                write_workbook(reports, workpaper)

            sheet = load_workbook(output, data_only=False)["Profit and Loss"]
            headers = [cell.value for row in sheet.iter_rows() for cell in row]
            self.assertNotIn("Tab 3 decision", headers)
            review_header = next(
                cell for row in sheet.iter_rows() for cell in row
                if cell.value == "Review note"
            )
            self.assertEqual(sheet.column_dimensions[review_header.column_letter].width, 64)
            self.assertTrue(sheet.cell(2, review_header.column).alignment.wrap_text)

    def test_balance_sheet_uses_the_same_compact_formula_summary_pattern_as_pl(self):
        with TemporaryDirectory() as folder:
            root = Path(folder)
            periods = ["2026"]
            reports = SimpleNamespace(
                pl_input=_source_report(root / "pl.xlsx", "P&L", periods, [("Net Profit", 100.0)]),
                bs_input=_source_report(root / "bs.xlsx", "BS", periods, [("Cash", 100.0)]),
                tax_depreciation_report=None,
                tax_depreciation_total=None,
                tax_depreciation_source=None,
            )
            output = root / "output.xlsx"
            with patch("v1.write_workbook.OUTPUT_PATH", output):
                write_workbook(reports, _workpaper(periods))

            sheet = load_workbook(output, data_only=False)["Balance Sheet"]
            values = [cell.value for row in sheet.iter_rows() for cell in row]
            formulas = [value for value in values if isinstance(value, str) and value.startswith("=SUMIF(")]
            self.assertIn("Balance Sheet ITR Totals", values)
            self.assertNotIn("Balance Sheet ITR Summary", values)
            self.assertTrue(formulas)

    def test_scalar_adjustment_and_loss_inputs_do_not_invent_period_values(self):
        periods = ["2026", "2025", "2024"]
        adjustment = {"description": "Reviewed adjustment", "amount": 125.0, "source": "Reviewer"}
        with patch.dict("v1.workpaper_builder.TAX_ADJUSTMENTS", {"add_back_7W": [adjustment]}, clear=True):
            add_rows, _, _, _ = _manual_adjustment_rows(periods)

        self.assertEqual(add_rows[0]["2026"], 125.0)
        self.assertEqual(add_rows[0]["2025"], 0.0)
        self.assertEqual(add_rows[0]["2024"], 0.0)
        loss_inputs = _build_carry_forward_losses_input(["2026"])
        self.assertEqual(loss_inputs["Period"].tolist(), ["2026"])
        self.assertTrue(loss_inputs["Opening losses"].isna().all())
        self.assertTrue(loss_inputs["Closing losses"].isna().all())


if __name__ == "__main__":
    unittest.main()
