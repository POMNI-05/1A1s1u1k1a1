from __future__ import annotations

import json
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from frontend.revision_editor import (
    RevisionEditorError,
    collect_revision_changes,
    export_revision,
    list_review_sheets,
    load_review_rows,
)


class RevisionEditorTests(unittest.TestCase):
    def _create_workbook(self, folder: Path) -> Path:
        path = folder / "source.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Profit and Loss"
        sheet.append(["Account", "2026", "ITR Ref", "Tab 3 decision", "Confidence", "Mapping reason", "Review note"])
        sheet.append(["Entertainment", 277.20, "Exp - 6S", "No use in Tab 3", "low", "Section fallback", "Confirm purpose"])
        sheet.append(["Bank fees", 100, "Exp - 6S", "No use in Tab 3", "high", "Rule matched", ""])
        workbook.save(path)
        workbook.close()
        return path

    def test_review_grid_is_detected_and_only_review_fields_can_change(self):
        with TemporaryDirectory() as folder:
            source = self._create_workbook(Path(folder))
            sheet = list_review_sheets(source)[0]
            original = load_review_rows(source, sheet)
            edited = [dict(row) for row in original]
            edited[0]["ITR Ref"] = "Review"
            edited[0]["Review note"] = "Invoice shows a private component."
            changes = collect_revision_changes(original, edited)

        self.assertEqual([change["field"] for change in changes], ["ITR Ref", "Review note"])
        self.assertEqual(changes[0]["old_value"], "Exp - 6S")
        self.assertEqual(changes[0]["new_value"], "Review")

    def test_export_creates_new_workbook_and_audit_without_touching_original(self):
        with TemporaryDirectory() as folder:
            source = self._create_workbook(Path(folder))
            sheet = list_review_sheets(source)[0]
            original = load_review_rows(source, sheet)
            edited = [dict(row) for row in original]
            edited[0]["ITR Ref"] = "Review"
            changes = collect_revision_changes(original, edited)
            exported = export_revision(
                source_workbook=source,
                sheet=sheet,
                changes=changes,
                reviewer="AB",
                revision_note="Pending invoice evidence.",
            )

            source_book = load_workbook(source, data_only=False)
            revision_book = load_workbook(exported.workbook_path, data_only=False)
            audit = json.loads(exported.audit_path.read_text(encoding="utf-8"))

        self.assertEqual(source_book["Profit and Loss"].cell(2, 3).value, "Exp - 6S")
        self.assertEqual(revision_book["Profit and Loss"].cell(2, 3).value, "Review")
        self.assertNotEqual(source, exported.workbook_path)
        self.assertEqual(audit["reviewer"], "AB")
        self.assertEqual(audit["changes"][0]["field"], "ITR Ref")
        source_book.close()
        revision_book.close()

    def test_formula_like_review_text_and_blank_itr_ref_are_rejected(self):
        rows = [
            {
                "Excel row": 2,
                "Account": "Entertainment",
                "ITR Ref": "Exp - 6S",
                "Confidence": "low",
                "Review note": "",
            }
        ]
        blank_ref = [dict(rows[0], **{"ITR Ref": ""})]
        formula_note = [dict(rows[0], **{"Review note": "=SUM(A1:A2)"})]

        with self.assertRaisesRegex(RevisionEditorError, "ITR Ref cannot be blank"):
            collect_revision_changes(rows, blank_ref)
        with self.assertRaisesRegex(RevisionEditorError, "formula character"):
            collect_revision_changes(rows, formula_note)
