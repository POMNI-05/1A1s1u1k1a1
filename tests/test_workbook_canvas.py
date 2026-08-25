from __future__ import annotations

import json
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from frontend.workbook_canvas import (
    _CANVAS_JS,
    export_manual_workbook_revision,
    load_workbook_canvas,
    merge_workbook_edits,
)


class WorkbookCanvasTests(unittest.TestCase):
    def test_canvas_javascript_preserves_clipboard_escape_sequences(self):
        """The inline JavaScript must not contain literal CR/LF regex characters."""

        self.assertIn(r"replace(/\r/g,'').split('\n').map(x=>x.split('\t'))", _CANVAS_JS)
        self.assertNotIn("\r", _CANVAS_JS)

    def _source_workbook(self, folder: Path) -> Path:
        path = folder / "source.xlsx"
        workbook = Workbook()
        workbook.active.title = "Profit and Loss"
        workbook.active["A1"] = "Revenue"
        workbook.active["B1"] = 100
        workbook.create_sheet("Tax Reconciliation")["A1"] = "=SUM('Profit and Loss'!B1)"
        workbook.save(path)
        workbook.close()
        return path

    def test_all_existing_cells_are_eligible_and_export_creates_a_manual_copy(self):
        with TemporaryDirectory() as folder:
            source = self._source_workbook(Path(folder))
            sheets = load_workbook_canvas(source)
            edits = merge_workbook_edits(
                sheets,
                {
                    "Profit and Loss::0::0": "Sales revenue",
                    "Profit and Loss::0::1": "125.50",
                    "Tax Reconciliation::0::0": "=SUM('Profit and Loss'!B1)*2",
                    "Unknown::0::0": "discarded",
                },
            )
            revision, audit_path, count = export_manual_workbook_revision(
                source_workbook=source,
                sheets=sheets,
                edits=edits,
            )
            source_book = load_workbook(source, data_only=False)
            revision_book = load_workbook(revision, data_only=False)
            audit = json.loads(audit_path.read_text(encoding="utf-8"))

        self.assertEqual(count, 3)
        self.assertEqual(source_book["Profit and Loss"]["A1"].value, "Revenue")
        self.assertEqual(source_book["Profit and Loss"]["B1"].value, 100)
        self.assertEqual(revision_book["Profit and Loss"]["A1"].value, "Sales revenue")
        self.assertEqual(revision_book["Profit and Loss"]["B1"].value, 125.5)
        self.assertEqual(revision_book["Tax Reconciliation"]["A1"].value, "=SUM('Profit and Loss'!B1)*2")
        self.assertEqual(audit["kind"], "unrestricted_manual_web_workbook_edit")
        source_book.close()
        revision_book.close()
