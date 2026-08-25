# v1/write_workbook.py
"""Write a reviewer-friendly Excel workpaper from validated report inputs.

This writer is input-source flexible:
- P&L and BS may come from separate workbooks;
- or from different sheets in one combined workbook.

Output rule:
- source P&L and Balance Sheet evidence stays on its own sheet;
- generated review columns are appended outside the copied source range;
- Tax Reconciliation contains only the reconciliation, not duplicated source reports;
- period columns come from validated data and are never padded to a fixed year count.
"""

from __future__ import annotations

import copy
import logging
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from .cleaner import ReportInput
    from .config import (
        OUTPUT_PATH,
        SELECTED_INCOME_YEAR,
        SHEET_BS_RAW,
        SHEET_PL_RAW,
        SHEET_RECONCILIATION,
    )
except ImportError:  # Direct-script compatibility.
    from cleaner import ReportInput
    from config import (
        OUTPUT_PATH,
        SELECTED_INCOME_YEAR,
        SHEET_BS_RAW,
        SHEET_PL_RAW,
        SHEET_RECONCILIATION,
    )

logger = logging.getLogger(__name__)

TITLE_FILL = PatternFill("solid", fgColor="B4C6E7")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill("solid", fgColor="FFFF00")
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")
RESULT_FILL = PatternFill("solid", fgColor="D9E2F3")
INPUT_FILL = PatternFill("solid", fgColor="E2F0D9")
APPROVED_ADJUSTMENT_FILL = PatternFill("solid", fgColor="E2F0D9")
APPROVED_DEDUCTION_FILL = PatternFill("solid", fgColor="D9EAF7")
PENDING_REVIEW_FILL = PatternFill("solid", fgColor="F4CCCC")

RED_FONT = Font(color="FF0000", bold=True, size=9)
TITLE_FONT = Font(bold=True, size=9)
HEADER_FONT = Font(bold=True, size=9)
BOLD_FONT = Font(bold=True, size=9)
NOTE_FONT = Font(italic=True, color="666666", size=9)
REVIEW_NOTE_FONT = Font(size=9)

THIN_BORDER = Border(bottom=Side(style="thin", color="BFBFBF"))
THICK_BORDER = Border(bottom=Side(style="medium", color="000000"))


def _font_with_size(font_obj, size: int = 12) -> Font:
    return Font(
        name=font_obj.name,
        sz=size,
        b=font_obj.b,
        i=font_obj.i,
        u=font_obj.u,
        strike=font_obj.strike,
        color=copy.copy(font_obj.color),
        vertAlign=font_obj.vertAlign,
        charset=font_obj.charset,
        family=font_obj.family,
        scheme=font_obj.scheme,
        outline=font_obj.outline,
        shadow=font_obj.shadow,
        condense=font_obj.condense,
        extend=font_obj.extend,
    )


def _force_font_size_9(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.font = _font_with_size(cell.font, 9)


def _safe(value):
    return None if pd.isna(value) else value


def _is_number(value) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _write_raw_df_to_sheet(wb: Workbook, raw_df: pd.DataFrame, title: str):
    ws = wb.create_sheet(title)
    for r_idx, row in enumerate(raw_df.itertuples(index=False, name=None), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, _safe(value))

    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    return ws


def _copy_raw_df_area(
    raw_df: pd.DataFrame,
    dst_ws,
    start_row: int,
    start_col: int,
) -> tuple[int, int]:
    for r_idx, row in enumerate(raw_df.itertuples(index=False, name=None), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            dst_ws.cell(r_idx, c_idx, _safe(value))

    last_row = start_row + max(len(raw_df), 1) - 1
    last_col = start_col + max(len(raw_df.columns), 1) - 1

    for col_idx in range(start_col, last_col + 1):
        dst_ws.column_dimensions[get_column_letter(col_idx)].width = 14

    return last_row, last_col

def _get_source_ws(report_input: ReportInput):
    src_wb = load_workbook(report_input.source_path, data_only=False)

    if report_input.sheet_name in src_wb.sheetnames:
        return src_wb, src_wb[report_input.sheet_name]

    try:
        idx = int(report_input.sheet_name)
        return src_wb, src_wb.worksheets[idx]
    except Exception as exc:
        raise KeyError(
            f"Could not find sheet {report_input.sheet_name!r} in {report_input.source_path}"
        ) from exc


def _copy_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy.copy(src.protection)


def _copy_cell(src, dst) -> None:
    dst.value = src.value
    _copy_style(src, dst)

    if src.hyperlink:
        dst._hyperlink = copy.copy(src.hyperlink)

    if src.comment:
        dst.comment = copy.copy(src.comment)


def _copy_cell_translated(src, dst, copy_formulas: bool = True) -> None:
    if copy_formulas and isinstance(src.value, str) and src.value.startswith("="):
        dst.value = Translator(
            src.value,
            origin=src.coordinate,
        ).translate_formula(dst.coordinate)
    else:
        dst.value = src.value

    _copy_style(src, dst)

    if src.hyperlink:
        dst._hyperlink = copy.copy(src.hyperlink)

    if src.comment:
        dst.comment = copy.copy(src.comment)


def _copy_merged_ranges(src_ws, dst_ws, row_offset: int = 0, col_offset: int = 0) -> None:
    for merged_range in src_ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds

        dst_ws.merge_cells(
            start_row=min_row + row_offset,
            start_column=min_col + col_offset,
            end_row=max_row + row_offset,
            end_column=max_col + col_offset,
        )

def _looks_like_generated_column(src_ws, col_idx: int) -> bool:
    """
    Detect generated helper columns from old output files.

    If a user accidentally uploads a prior generated workbook, we do not want
    the copied Profit and Loss / Balance Sheet evidence tabs to include:
    - ITR Label
    - ITR Ref
    - Review note
    - ITR Totals
    - yellow summary blocks
    """
    values = []

    for row_idx in range(1, min(src_ws.max_row, 40) + 1):
        value = src_ws.cell(row_idx, col_idx).value
        if value is None:
            continue
        values.append(str(value).strip().lower())

    generated_markers = [
        "itr label",
        "itr ref",
        "itr totals",
        "review note",
        "tax reconciliation",
    ]

    # Only confirmed header values identify generated columns. Genuine account
    # columns naturally contain Total Income / Total Expenses / Net Profit.
    return any(value in generated_markers for value in values[:10])

def _copy_sheet_to_workbook(report_input: ReportInput, dst_wb: Workbook, title: str):
    """
    Copy selected original source sheet into output workbook.

    This version keeps the raw evidence sheet clean. If the source file is
    accidentally a previous generated output workbook, it skips generated
    helper columns such as ITR Label / ITR Totals / Review note.
    """
    _, src_ws = _get_source_ws(report_input)
    dst_ws = dst_wb.create_sheet(title)

    allowed_cols = [
        col_idx
        for col_idx in range(1, src_ws.max_column + 1)
        if not _looks_like_generated_column(src_ws, col_idx)
    ]

    col_map = {
        old_col_idx: new_col_idx
        for new_col_idx, old_col_idx in enumerate(allowed_cols, start=1)
    }

    for row in src_ws.iter_rows():
        for src_cell in row:
            if src_cell.column not in col_map:
                continue

            dst_cell = dst_ws.cell(src_cell.row, col_map[src_cell.column])
            _copy_cell(src_cell, dst_cell)

    for old_col_idx, new_col_idx in col_map.items():
        src_letter = get_column_letter(old_col_idx)
        dst_letter = get_column_letter(new_col_idx)
        dst_ws.column_dimensions[dst_letter].width = (
            src_ws.column_dimensions[src_letter].width or 12
        )

    for row_idx, dim in src_ws.row_dimensions.items():
        if dim.height is not None:
            dst_ws.row_dimensions[row_idx].height = dim.height

    # Do not copy merged ranges here, because generated columns may be skipped
    # and old merged ranges can become invalid. Raw values/styles are enough.
    dst_ws.freeze_panes = src_ws.freeze_panes
    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines

    return dst_ws


def _copy_report_area(
    report_input: ReportInput,
    dst_ws,
    start_row: int,
    start_col: int,
    copy_formulas: bool = True,
) -> tuple[int, int]:
    """Copy selected original report sheet into the reconciliation sheet."""
    _, src_ws = _get_source_ws(report_input)

    row_offset = start_row - 1
    col_offset = start_col - 1

    for row in src_ws.iter_rows():
        for src_cell in row:
            dst_cell = dst_ws.cell(
                row=start_row + src_cell.row - 1,
                column=start_col + src_cell.column - 1,
            )
            _copy_cell_translated(src_cell, dst_cell, copy_formulas=copy_formulas)

    for col_idx in range(1, src_ws.max_column + 1):
        src_letter = get_column_letter(col_idx)
        dst_letter = get_column_letter(start_col + col_idx - 1)
        dst_ws.column_dimensions[dst_letter].width = (
            src_ws.column_dimensions[src_letter].width or 12
        )

    # Preserve source row heights only. Do not auto-resize row height.
    for row_idx, dim in src_ws.row_dimensions.items():
        if dim.height is not None:
            dst_ws.row_dimensions[start_row + row_idx - 1].height = dim.height

    _copy_merged_ranges(src_ws, dst_ws, row_offset=row_offset, col_offset=col_offset)

    return start_row + src_ws.max_row - 1, start_col + src_ws.max_column - 1

def should_highlight_mapping(mapping: dict, report_type: str = "") -> bool:
    itr_ref = str(mapping.get("ITR Ref", "") or "").strip()
    treatment = str(mapping.get("Treatment", "") or "").lower().strip()
    confidence = str(mapping.get("Confidence", "") or "").lower().strip()

    if itr_ref == "Review":
        return True

    if confidence == "low":
        return True

    if report_type == "balance_sheet" and treatment == "support_only":
        return False

    if treatment == "review_only":
        return True

    return False

def _write_side_labels(
    ws,
    labelled_df: pd.DataFrame,
    source_start_row: int,
    itr_col: int,
    confidence_col: int,
    reason_col: int,
    review_col: int,
    report_type: str = "",
) -> None:
    source_rows = []
    if labelled_df is not None and not labelled_df.empty and "Source Row" in labelled_df:
        source_rows = [
            int(value)
            for value in labelled_df["Source Row"].dropna().tolist()
            if int(value) >= source_start_row
        ]

    header_row = max(source_start_row, min(source_rows) - 1) if source_rows else source_start_row
    headers = {
        itr_col: "ITR Ref",
        confidence_col: "Confidence",
        reason_col: "Mapping reason",
        review_col: "Review note",
    }

    for col_idx, header in headers.items():
        ws.cell(header_row, col_idx, header)

    for col_idx in headers:
        cell = ws.cell(header_row, col_idx)
        cell.font = RED_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=False)

    if labelled_df is None or labelled_df.empty:
        return

    for _, row in labelled_df.iterrows():
        source_row = row.get("Source Row")

        if pd.isna(source_row):
            continue

        row_type = str(row.get("Row Type", "") or "").lower().strip()
        if row_type not in {"account", "total"}:
            continue

        itr_ref = str(row.get("ITR Ref", "") or "").strip()
        review_note = str(row.get("Review Note", "") or "").strip()
        label_reason = str(row.get("Label Reason", "") or "").strip()
        confidence = str(row.get("Confidence", "") or "").lower().strip()

        if not itr_ref and not review_note and not label_reason:
            continue

        excel_row = source_start_row + int(source_row) - 1

        visible_note = review_note

        label_cell = ws.cell(excel_row, itr_col, itr_ref)
        confidence_cell = ws.cell(excel_row, confidence_col, confidence or None)
        reason_cell = ws.cell(excel_row, reason_col, label_reason or None)
        note_cell = ws.cell(excel_row, review_col, visible_note)

        label_cell.font = RED_FONT if itr_ref else REVIEW_NOTE_FONT
        label_cell.alignment = Alignment(vertical="top", wrap_text=False)

        for cell in (confidence_cell, reason_cell, note_cell):
            cell.font = REVIEW_NOTE_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        if should_highlight_mapping(row.to_dict(), report_type=report_type):
            for cell in (label_cell, confidence_cell, reason_cell, note_cell):
                cell.fill = REVIEW_FILL

def _write_tax_reconciliation_table(
    ws,
    df: pd.DataFrame,
    title: str,
    start_row: int,
    start_col: int,
) -> tuple[int, int]:
    display_cols = [c for c in df.columns if c != "Line Type"]
    last_col = start_col + len(display_cols) - 1

    for col in range(start_col, last_col + 1):
        ws.cell(start_row, col).fill = TITLE_FILL
        ws.cell(start_row, col).font = TITLE_FONT

    ws.cell(start_row, start_col, title)

    header_row = start_row + 1

    period_cols = [
        col
        for col in display_cols
        if col not in {"Description", "ITR Ref", "Tax return code", "Review note"}
    ]
    period_excel_cols = {
        col_name: start_col + display_cols.index(col_name)
        for col_name in period_cols
    }
    base_row_idx: int | None = None
    add_heading_row_idx: int | None = None
    add_total_row_idx: int | None = None
    subtract_heading_row_idx: int | None = None
    subtract_total_row_idx: int | None = None
    result_row_idx: int | None = None

    for idx, col_name in enumerate(display_cols):
        display_name = col_name
        if col_name in period_cols:
            year_match = re.search(r"\b(20\d{2})\b", str(col_name))
            if year_match:
                display_name = f"Year Ended 30 June {year_match.group(1)}"

        cell = ws.cell(header_row, start_col + idx, display_name)
        cell.fill = TITLE_FILL if str(SELECTED_INCOME_YEAR) in str(col_name) else HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=False)

    for r_idx, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        line_type = str(row.get("Line Type", "")).lower()
        description = str(row.get("Description", "") or "")

        if description == "Accounting profit/(loss) — Item 6T":
            base_row_idx = r_idx
        elif line_type == "add_heading":
            add_heading_row_idx = r_idx
        elif description == "Total ADD":
            add_total_row_idx = r_idx
        elif line_type == "subtract_heading":
            subtract_heading_row_idx = r_idx
        elif description == "Total SUBTRACT":
            subtract_total_row_idx = r_idx
        elif line_type == "result" and "Item 7T" in description:
            result_row_idx = r_idx

        for c_idx, col_name in enumerate(display_cols):
            value = _safe(row[col_name])
            cell = ws.cell(r_idx, start_col + c_idx, value)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=str(col_name).strip().lower() in {"detail", "review note", "reason"},
            )

            if _is_number(value):
                cell.number_format = '$#,##0.00;($#,##0.00);-'

            if col_name == "ITR Ref" and str(value or "").strip():
                cell.font = RED_FONT

            if col_name in period_cols:
                cell.fill = (
                    TITLE_FILL
                    if str(SELECTED_INCOME_YEAR) in str(col_name)
                    else HEADER_FILL
                )

        if line_type == "heading":
            for c in range(start_col, last_col + 1):
                ws.cell(r_idx, c).fill = SECTION_FILL
                ws.cell(r_idx, c).font = BOLD_FONT

        elif line_type in {"add_heading", "subtract_heading"}:
            # Direction must be immediately visible without turning the whole
            # review calculation into a coloured warning block.
            ws.cell(r_idx, start_col).font = RED_FONT

        elif line_type == "placeholder":
            for c in range(start_col, last_col + 1):
                ws.cell(r_idx, c).font = NOTE_FONT

        elif line_type == "review":
            for c in range(start_col, last_col + 1):
                ws.cell(r_idx, c).fill = REVIEW_FILL
            ws.cell(r_idx, start_col).font = BOLD_FONT

        elif line_type == "subtotal":
            for c in range(start_col, last_col + 1):
                ws.cell(r_idx, c).font = BOLD_FONT
                ws.cell(r_idx, c).border = THIN_BORDER

        elif line_type == "result":
            for c in range(start_col, last_col + 1):
                ws.cell(r_idx, c).font = BOLD_FONT
                ws.cell(r_idx, c).border = THICK_BORDER
            ws.cell(r_idx, start_col).fill = RESULT_FILL

        elif line_type == "note":
            for c in range(start_col, last_col + 1):
                ws.cell(r_idx, c).font = NOTE_FONT

    # Keep the deterministic Python calculation as the generation-time
    # control, but make the visible Tab 3 bridge live in Excel.  An accountant
    # can now amend an adjustment and immediately see the totals and 7T move.
    for period, col_idx in period_excel_cols.items():
        letter = get_column_letter(col_idx)

        if add_total_row_idx is not None and add_heading_row_idx is not None:
            ws.cell(add_total_row_idx, col_idx).value = (
                f"=SUM({letter}{add_heading_row_idx + 1}:{letter}{add_total_row_idx - 1})"
            )

        if subtract_total_row_idx is not None and subtract_heading_row_idx is not None:
            ws.cell(subtract_total_row_idx, col_idx).value = (
                f"=SUM({letter}{subtract_heading_row_idx + 1}:{letter}{subtract_total_row_idx - 1})"
            )

        if result_row_idx is not None and base_row_idx is not None:
            add_term = f"{letter}{add_total_row_idx}" if add_total_row_idx is not None else "0"
            subtract_term = (
                f"{letter}{subtract_total_row_idx}"
                if subtract_total_row_idx is not None
                else "0"
            )
            ws.cell(result_row_idx, col_idx).value = (
                f"={letter}{base_row_idx}+{add_term}-{subtract_term}"
            )

    for col in range(start_col, last_col + 1):
        header = str(ws.cell(header_row, col).value or "")
        letter = get_column_letter(col)

        if header == "Description":
            ws.column_dimensions[letter].width = 34
        elif header == "Review note":
            ws.column_dimensions[letter].width = 60
        elif re.search(r"20\d{2}|30 June|30 Jun", header):
            ws.column_dimensions[letter].width = 22
        else:
            ws.column_dimensions[letter].width = 12

    return header_row + len(df), last_col
PL_INCOME_SUMMARY_LABELS = [
    "Inc - 6A",
    "Inc - 6B",
    "Inc - 6C",
    "Inc - 6D",
    "Inc - 6E",
    "Inc - 6X",
    "Inc - 6F",
    "Inc - 6G",
    "Inc - 6H",
    "Inc - 6I",
    "Inc - 6Q",
    "Inc - 6J",
    "Inc - 6R",
]

PL_EXPENSE_SUMMARY_LABELS = [
    "Exp - 6A",
    "Exp - 6C",
    "Exp - 6D",
    "Exp - 6E",
    "Exp - 6F",
    "Exp - 6I",
    "Exp - 6H",
    "Exp - 6V",
    "Exp - 6J",
    "Exp - 6U",
    "Exp - 6W",
    "Exp - 6X",
    "Exp - 6Y",
    "Exp - 6Z",
    "Exp - 6G",
    "Exp - 6S",
]

def _find_main_amount_col(ws, start_row: int, last_row: int, last_col: int) -> int:
    """
    Find the main numeric amount column in the copied report block.

    Usually this is column B, but this is safer than hard-coding B.
    """
    # A report title often contains the selected year in column A (for example
    # "For the year ended 30 June 2025").  It is context, not an amount
    # column.  First find the actual Account/Description header row and choose
    # the selected period only from that same row.
    header_rows_found = False
    for row_idx in range(start_row, min(last_row, start_row + 12) + 1):
        values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, last_col + 1)]
        has_account_header = any(
            str(value or "").strip().lower()
            in {"account", "account name", "account label", "description"}
            for value in values
        )
        if not has_account_header:
            continue

        header_rows_found = True
        year_matches = [
            col_idx
            for col_idx, value in enumerate(values, start=1)
            if value is not None and str(SELECTED_INCOME_YEAR) in str(value)
        ]
        if len(year_matches) == 1:
            return year_matches[0]
        if len(year_matches) > 1:
            raise ValueError(
                f"PERIOD-001: income year {SELECTED_INCOME_YEAR} appears in multiple "
                f"amount headers: {year_matches}"
            )

    if header_rows_found:
        raise ValueError(
            f"PERIOD-001: income year {SELECTED_INCOME_YEAR} is not present in the "
            "Account/Description header row."
        )

    # A generic one-period export may omit a year header.  It is safe only
    # when exactly one numeric candidate exists; otherwise do not guess.
    numeric_cols = []
    for col_idx in range(1, last_col + 1):
        for row_idx in range(start_row, last_row + 1):
            value = ws.cell(row_idx, col_idx).value
            if _is_number(value):
                numeric_cols.append(col_idx)
                break

    if len(numeric_cols) == 1:
        return numeric_cols[0]

    raise ValueError(
        f"PERIOD-001: income year {SELECTED_INCOME_YEAR} is not uniquely present in report columns; "
        f"numeric candidates={numeric_cols}"
    )


def _write_pl_formula_summary_table(
    ws,
    title: str,
    start_row: int,
    start_col: int,
    label_col: int,
    amount_col: int,
    data_start_row: int,
    data_last_row: int,
) -> tuple[int, int]:
    """
    Write P&L ITR totals using Excel formulas.

    Formula pattern:
    =SUMIF($<label_col>$start:$<label_col>$end, <summary_label_cell>, $<amount_col>$start:$<amount_col>$end)
    """
    label_letter = get_column_letter(label_col)
    amount_letter = get_column_letter(amount_col)

    label_range = f"${label_letter}${data_start_row}:${label_letter}${data_last_row}"
    amount_range = f"${amount_letter}${data_start_row}:${amount_letter}${data_last_row}"

    label_out_col = start_col
    amount_out_col = start_col + 1

    # Title/header
    ws.cell(start_row, label_out_col, title)
    ws.cell(start_row, label_out_col).fill = SECTION_FILL
    ws.cell(start_row, label_out_col).font = BOLD_FONT

    ws.cell(start_row, amount_out_col, "Amount")
    ws.cell(start_row, amount_out_col).fill = SECTION_FILL
    ws.cell(start_row, amount_out_col).font = BOLD_FONT

    current_row = start_row + 2

    income_amount_cells = []
    expense_amount_cells = []

    # Income section
    ws.cell(current_row, label_out_col, "Income")
    ws.cell(current_row, label_out_col).font = BOLD_FONT
    current_row += 1

    for label in PL_INCOME_SUMMARY_LABELS:
        label_cell = ws.cell(current_row, label_out_col, label)
        amount_cell = ws.cell(current_row, amount_out_col)

        amount_cell.value = f"=SUMIF({label_range},{label_cell.coordinate},{amount_range})"
        amount_cell.number_format = '$#,##0.00;($#,##0.00);-'

        income_amount_cells.append(amount_cell.coordinate)
        current_row += 1

    total_income_row = current_row
    ws.cell(total_income_row, label_out_col, "TOTAL INCOME")
    ws.cell(total_income_row, label_out_col).font = BOLD_FONT
    ws.cell(total_income_row, amount_out_col, f"=SUM({','.join(income_amount_cells)})")
    ws.cell(total_income_row, amount_out_col).number_format = '$#,##0.00;($#,##0.00);-'
    ws.cell(total_income_row, amount_out_col).font = BOLD_FONT

    current_row += 3

    # Expense section
    ws.cell(current_row, label_out_col, "Expenses")
    ws.cell(current_row, label_out_col).font = BOLD_FONT
    current_row += 1

    for label in PL_EXPENSE_SUMMARY_LABELS:
        label_cell = ws.cell(current_row, label_out_col, label)
        amount_cell = ws.cell(current_row, amount_out_col)

        amount_cell.value = f"=SUMIF({label_range},{label_cell.coordinate},{amount_range})"
        amount_cell.number_format = '$#,##0.00;($#,##0.00);-'

        expense_amount_cells.append(amount_cell.coordinate)
        current_row += 1

    total_expense_row = current_row
    ws.cell(total_expense_row, label_out_col, "TOTAL EXPENSES")
    ws.cell(total_expense_row, label_out_col).font = BOLD_FONT
    ws.cell(total_expense_row, amount_out_col, f"=SUM({','.join(expense_amount_cells)})")
    ws.cell(total_expense_row, amount_out_col).number_format = '$#,##0.00;($#,##0.00);-'
    ws.cell(total_expense_row, amount_out_col).font = BOLD_FONT

    current_row += 2

    # Pre-tax profit
    pre_tax_row = current_row
    ws.cell(pre_tax_row, label_out_col, "PRE TAX PROFIT/(LOSS)")
    ws.cell(pre_tax_row, label_out_col).font = BOLD_FONT
    ws.cell(
        pre_tax_row,
        amount_out_col,
        f"={ws.cell(total_income_row, amount_out_col).coordinate}-{ws.cell(total_expense_row, amount_out_col).coordinate}",
    )
    ws.cell(pre_tax_row, amount_out_col).number_format = '$#,##0.00;($#,##0.00);-'
    ws.cell(pre_tax_row, amount_out_col).font = BOLD_FONT

    # Format yellow block
    for row_idx in range(start_row, pre_tax_row + 1):
        for col_idx in range(label_out_col, amount_out_col + 1):
            ws.cell(row_idx, col_idx).fill = SECTION_FILL

    ws.column_dimensions[get_column_letter(label_out_col)].width = 18
    ws.column_dimensions[get_column_letter(amount_out_col)].width = 16

    return pre_tax_row, amount_out_col


def _balance_sheet_summary_labels(labelled_df: pd.DataFrame) -> list[str]:
    """Return the filing/support references worth showing in a compact total."""

    if labelled_df is None or labelled_df.empty or "ITR Ref" not in labelled_df.columns:
        return []

    rows = labelled_df.copy()
    if "Row Type" in rows.columns:
        rows = rows[rows["Row Type"].astype(str).str.lower().isin({"account", "total"})]
    if "Treatment" in rows.columns:
        rows = rows[~rows["Treatment"].astype(str).str.lower().eq("support_only")]

    refs = {
        str(value).strip()
        for value in rows["ITR Ref"].tolist()
        if str(value).strip() and str(value).strip() != "Review"
    }
    return sorted(refs)


def _write_balance_sheet_formula_summary_table(
    ws,
    start_row: int,
    start_col: int,
    label_col: int,
    amount_col: int,
    data_start_row: int,
    data_last_row: int,
    labelled_df: pd.DataFrame,
) -> tuple[int, int]:
    """Write a concise Item 8 review total beside the Balance Sheet evidence."""

    title = "Balance Sheet ITR Totals"
    labels = _balance_sheet_summary_labels(labelled_df)
    ws.cell(start_row, start_col, title)
    ws.cell(start_row, start_col).fill = SECTION_FILL
    ws.cell(start_row, start_col).font = BOLD_FONT
    ws.cell(start_row, start_col + 1, "Amount")
    ws.cell(start_row, start_col + 1).fill = SECTION_FILL
    ws.cell(start_row, start_col + 1).font = BOLD_FONT

    if not labels:
        ws.cell(start_row + 1, start_col, "No filing/support labels detected")
        ws.cell(start_row + 1, start_col).font = NOTE_FONT
        return start_row + 1, start_col + 1

    label_letter = get_column_letter(label_col)
    amount_letter = get_column_letter(amount_col)
    label_range = f"${label_letter}${data_start_row}:${label_letter}${data_last_row}"
    amount_range = f"${amount_letter}${data_start_row}:${amount_letter}${data_last_row}"

    for row_offset, label in enumerate(labels, start=2):
        label_cell = ws.cell(start_row + row_offset, start_col, label)
        amount_cell = ws.cell(start_row + row_offset, start_col + 1)
        amount_cell.value = f"=SUMIF({label_range},{label_cell.coordinate},{amount_range})"
        amount_cell.number_format = '$#,##0.00;($#,##0.00);-'

    for row_idx in range(start_row, start_row + len(labels) + 2):
        for col_idx in range(start_col, start_col + 2):
            ws.cell(row_idx, col_idx).fill = SECTION_FILL

    ws.column_dimensions[get_column_letter(start_col)].width = 24
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 16
    return start_row + len(labels) + 1, start_col + 1

def _write_label_summary_table(
    ws,
    df: pd.DataFrame,
    title: str,
    start_row: int,
    start_col: int,
) -> tuple[int, int]:
    """
    Write P&L or BS label summary table.

    This is the system-generated equivalent of the old yellow SUMIF table,
    but safer because it comes from labelled dataframe results rather than
    hard-coded Excel column letters like E:E or F:F.
    """
    if df is None or df.empty:
        ws.cell(start_row, start_col, title)
        ws.cell(start_row + 1, start_col, "No labels detected")
        ws.cell(start_row, start_col).fill = TITLE_FILL
        ws.cell(start_row, start_col).font = TITLE_FONT
        return start_row + 1, start_col

    display_df = df.copy()

    last_col = start_col + len(display_df.columns) - 1

    for col in range(start_col, last_col + 1):
        ws.cell(start_row, col).fill = TITLE_FILL
        ws.cell(start_row, col).font = TITLE_FONT

    ws.cell(start_row, start_col, title)

    header_row = start_row + 1

    for idx, col_name in enumerate(display_df.columns):
        cell = ws.cell(header_row, start_col + idx, col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=False)

    for r_idx, (_, row) in enumerate(display_df.iterrows(), start=header_row + 1):
        for c_idx, col_name in enumerate(display_df.columns):
            value = _safe(row[col_name])
            cell = ws.cell(r_idx, start_col + c_idx, value)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=str(col_name).strip().lower()
                in {"review note", "reason", "label reason", "detail"},
            )
            cell.border = THIN_BORDER

            if col_name == "Amount" and _is_number(value):
                cell.number_format = '$#,##0.00;($#,##0.00);-'

            if col_name == "ITR Ref" and str(value or "").strip():
                cell.font = RED_FONT

            confidence = str(row.get("Confidence", "") or "").lower()
            treatment = str(row.get("Treatment", "") or "").lower()

            if confidence in {"low", "medium"} or treatment == "review_only":
                cell.fill = REVIEW_FILL

    for c_idx, col_name in enumerate(display_df.columns, start=start_col):
        letter = get_column_letter(c_idx)

        if col_name in {"ITR Label", "Review Note"}:
            ws.column_dimensions[letter].width = 34
        elif col_name == "Source Rows":
            ws.column_dimensions[letter].width = 18
        elif col_name == "Amount":
            ws.column_dimensions[letter].width = 16
        else:
            ws.column_dimensions[letter].width = 14

    return header_row + len(display_df), last_col

def _write_simple_table(
    ws,
    df: pd.DataFrame,
    title: str,
    start_row: int,
    start_col: int,
    input_table: bool = False,
) -> tuple[int, int]:
    if df is None or df.empty:
        title_cell = ws.cell(start_row, start_col, title)
        title_cell.fill = TITLE_FILL
        title_cell.font = TITLE_FONT
        return start_row, start_col

    last_col = start_col + len(df.columns) - 1

    for col in range(start_col, last_col + 1):
        cell = ws.cell(start_row, col)
        cell.fill = TITLE_FILL
        cell.font = TITLE_FONT

    ws.cell(start_row, start_col, title)

    header_row = start_row + 1

    for idx, col_name in enumerate(df.columns):
        cell = ws.cell(header_row, start_col + idx, str(col_name))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=False)

    for r_idx, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        for c_idx, col_name in enumerate(df.columns):
            value = _safe(row[col_name])
            cell = ws.cell(r_idx, start_col + c_idx, value)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = THIN_BORDER

            if input_table and value is None:
                cell.fill = INPUT_FILL

            if _is_number(value):
                cell.number_format = '$#,##0.00;($#,##0.00);-'

    for c_idx, col_name in enumerate(df.columns, start=start_col):
        letter = get_column_letter(c_idx)
        lower = str(col_name).strip().lower()

        if lower in {"review note", "reason", "label reason", "detail"}:
            ws.column_dimensions[letter].width = 60
        elif lower == "check":
            ws.column_dimensions[letter].width = 32
        elif lower == "status":
            ws.column_dimensions[letter].width = 26
        elif lower in {"description", "account", "workpaper label", "itr label"}:
            ws.column_dimensions[letter].width = 30
        elif lower == "source rows":
            ws.column_dimensions[letter].width = 16
        else:
            ws.column_dimensions[letter].width = 14

    return header_row + len(df), last_col

def _write_bs_checks(ws, bs_checks: pd.DataFrame, start_row: int, start_col: int = 1) -> int:
    if bs_checks is None or bs_checks.empty:
        return start_row

    end_row, _ = _write_simple_table(ws, bs_checks, "Balance Sheet Test Checks", start_row, start_col)

    for row in range(start_row + 2, end_row + 1):
        for col in range(start_col, start_col + len(bs_checks.columns)):
            value = ws.cell(row, col).value
            if _is_number(value) and abs(value) > 1:
                ws.cell(row, col).fill = REVIEW_FILL

    return end_row


def _format_sheet(ws, raw_block_last_col: int) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True

    for col_idx in range(raw_block_last_col + 1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        current = ws.column_dimensions[letter].width or 10
        ws.column_dimensions[letter].width = min(max(current, 10), 40)

    for col_idx in range(raw_block_last_col + 2, raw_block_last_col + 4):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22


def _append_review_output(
    ws,
    labelled_df: pd.DataFrame,
    label_summary: pd.DataFrame,
    report_type: str,
) -> None:
    """Append generated review fields without changing the copied source range."""
    source_last_col = ws.max_column
    itr_col = source_last_col + 2
    confidence_col = itr_col + 1
    reason_col = confidence_col + 1
    review_col = reason_col + 1

    widths = {
        itr_col: 14,
        confidence_col: 12,
        reason_col: 44,
        review_col: 64,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    _write_side_labels(
        ws,
        labelled_df,
        1,
        itr_col,
        confidence_col,
        reason_col,
        review_col,
        report_type=report_type,
    )

    summary_col = review_col + 2
    if report_type == "profit_and_loss":
        amount_col = _find_main_amount_col(ws, 1, ws.max_row, source_last_col)
        _write_pl_formula_summary_table(
            ws,
            title="ITR Totals",
            start_row=1,
            start_col=summary_col,
            label_col=itr_col,
            amount_col=amount_col,
            data_start_row=1,
            data_last_row=ws.max_row,
        )
    else:
        amount_col = _find_main_amount_col(ws, 1, ws.max_row, source_last_col)
        _write_balance_sheet_formula_summary_table(
            ws,
            1,
            summary_col,
            itr_col,
            amount_col,
            1,
            ws.max_row,
            labelled_df,
        )

    ws.freeze_panes = ws.freeze_panes or "A2"


def _write_inputs_sheet(wb: Workbook, workpaper) -> None:
    tables: list[tuple[str, pd.DataFrame, bool]] = []

    if workpaper.carry_forward_losses is not None and not workpaper.carry_forward_losses.empty:
        tables.append(("Carry Forward Losses", workpaper.carry_forward_losses, True))
    if workpaper.rd_breakdown is not None and not workpaper.rd_breakdown.empty:
        tables.append(("R&D Breakdown", workpaper.rd_breakdown, True))

    for title, table in getattr(workpaper, "support_tables", {}).items():
        if table is not None and not table.empty:
            tables.append((title, table, True))

    proposed = getattr(workpaper, "proposed_adjustments", None)
    if proposed is not None and not proposed.empty:
        tables.append(("Proposed Tax Adjustments - Not Posted Unless Approved", proposed, False))

    if not tables:
        return

    ws = wb.create_sheet("Inputs & Overrides")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    row = 1
    for title, table, is_input in tables:
        row, _ = _write_simple_table(ws, table, title, row, 1, input_table=is_input)
        row += 3


def _write_checks_sheet(wb: Workbook, reports, workpaper) -> None:
    ws = wb.create_sheet("Checks")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    periods = [
        col
        for col in workpaper.tax_reconciliation.columns
        if col not in {
            "Line Type",
            "Description",
            "ITR Ref",
            "Tax return code",
            "Review note",
        }
    ]
    output_checks = pd.DataFrame(
        [
            {"Check": "Requested income year", "Status": "OK", "Detail": str(SELECTED_INCOME_YEAR)},
            {
                "Check": "Rendered reconciliation periods",
                "Status": "OK" if periods else "NOT TESTED - PERIOD-001",
                "Detail": ", ".join(map(str, periods)) if periods else "No validated periods",
            },
            {"Check": "Missing periods generated", "Status": "OK", "Detail": "No - output uses only validated source periods"},
            {
                "Check": "Fixed Assets evidence",
                "Status": "OK" if reports.tax_depreciation_report is not None else "NOT PROVIDED",
                "Detail": reports.tax_depreciation_source or "No validated depreciation schedule",
            },
        ]
    )
    row, _ = _write_simple_table(ws, output_checks, "Workbook Output Checks", 1, 1)

    if workpaper.bs_checks is not None and not workpaper.bs_checks.empty:
        row, _ = _write_simple_table(ws, workpaper.bs_checks, "Balance Sheet Test Checks", row + 3, 1)

    tax_reconciliation_checks = getattr(workpaper, "tax_reconciliation_review_checks", None)
    if tax_reconciliation_checks is not None and not tax_reconciliation_checks.empty:
        row, _ = _write_simple_table(
            ws,
            tax_reconciliation_checks,
            "Tax Reconciliation Review Checks",
            row + 3,
            1,
        )

    review_items = getattr(workpaper, "review_items", None)
    if review_items is not None and not review_items.empty:
        _write_simple_table(ws, review_items, "Review Items", row + 3, 1)


def write_workbook(reports, workpaper) -> None:
    logger.info("Writing workbook to %s", OUTPUT_PATH)

    wb = Workbook()
    wb.remove(wb.active)

    pl_ws = _copy_sheet_to_workbook(reports.pl_input, wb, SHEET_PL_RAW)
    bs_ws = _copy_sheet_to_workbook(reports.bs_input, wb, SHEET_BS_RAW)
    _append_review_output(pl_ws, workpaper.labelled_pl, workpaper.pl_label_summary, "profit_and_loss")
    _append_review_output(bs_ws, workpaper.labelled_bs, workpaper.bs_label_summary, "balance_sheet")

    # Uploaded source evidence always comes first.  A supplied tax-depreciation
    # schedule is copied before the generated calculation/review tabs so a
    # reviewer reads source workpapers before the derived output.
    if reports.tax_depreciation_report is not None and reports.tax_depreciation_total is not None:
        _copy_sheet_to_workbook(reports.tax_depreciation_report, wb, "Fixed Assets")

    ws = wb.create_sheet(SHEET_RECONCILIATION)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    _, tax_last_col = _write_tax_reconciliation_table(
        ws,
        workpaper.tax_reconciliation,
        "Income Tax Reconciliation",
        1,
        1,
    )
    _format_sheet(ws, tax_last_col)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    tax_calculation = getattr(workpaper, "tax_calculation", None)
    if tax_calculation is not None and not tax_calculation.empty:
        tax_ws = wb.create_sheet("Tax Calculation")
        tax_ws.sheet_view.showGridLines = False
        tax_ws.freeze_panes = "A3"
        _, calculation_last_col = _write_tax_reconciliation_table(
            tax_ws,
            tax_calculation,
            "Company Tax Calculation — outside Item 7",
            1,
            1,
        )
        _format_sheet(tax_ws, calculation_last_col)
        tax_ws.sheet_view.showGridLines = False
        tax_ws.freeze_panes = "A3"

    _write_inputs_sheet(wb, workpaper)
    _write_checks_sheet(wb, reports, workpaper)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    Path(OUTPUT_PATH).parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)

    logger.info("Workbook saved: %s", OUTPUT_PATH)
