# v1_selenium/utils.py

import os
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from config import LOG_DIR, OUTPUT_PATH, SHEET_RECONCILIATION


def setup_logging():
    """
    Set up logging to both console and a timestamped log file.
    Call once at the top of main.py.
    """
    os.makedirs(LOG_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file  = os.path.join(LOG_DIR, f"run_{timestamp}.log")

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler(),          # also print to terminal
        ]
    )
    logging.info(f"Logging started — {log_file}")


def ensure_dirs():
    """Create data/, output/, logs/ if they don't exist."""
    from config import DOWNLOAD_DIR, OUTPUT_DIR, LOG_DIR
    for d in [DOWNLOAD_DIR, OUTPUT_DIR, LOG_DIR]:
        os.makedirs(d, exist_ok=True)


def format_workbook():
    """
    Apply formatting to the final output workbook:
    - Bold + grey header row on all sheets
    - Auto column widths
    - Freeze top row
    - Red fill on any cell containing '⚠'
    """
    wb = load_workbook(OUTPUT_PATH)

    HEADER_FILL  = PatternFill("solid", fgColor="D9D9D9")   # light grey
    WARNING_FILL = PatternFill("solid", fgColor="FFCCCC")   # light red
    HEADER_FONT  = Font(bold=True)

    for ws in wb.worksheets:
        # Bold + fill header row
        for cell in ws[1]:
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(horizontal="left")

        # Freeze top row
        ws.freeze_panes = "A2"

        # Auto column width
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value),
                default=10
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

        # Red highlight on warning rows (reconciliation sheet only)
        if ws.title == SHEET_RECONCILIATION:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and "⚠" in str(cell.value):
                        for r in row:
                            r.fill = WARNING_FILL

    wb.save(OUTPUT_PATH)
    logging.getLogger(__name__).info("✓ Workbook formatted and saved.")