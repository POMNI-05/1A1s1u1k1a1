# v1_selenium/write_workbook.py

from __future__ import annotations

import copy
import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_PATH, PL_RAW_PATH, BS_RAW_PATH

logger = logging.getLogger(__name__)

TITLE_FILL = PatternFill("solid", fgColor="B4C6E7")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")
RESULT_FILL = PatternFill("solid", fgColor="D9EAD3")

RED_FONT = Font(color="FF0000", bold=True)
TITLE_FONT = Font(bold=True)
HEADER_FONT = Font(bold=True)
BOLD_FONT = Font(bold=True)

THIN_BORDER = Border(bottom=Side(style="thin", color="808080"))
THICK_BORDER = Border(bottom=Side(style="medium", color="000000"))


def _copy_cell(src_cell, dst_cell):
    dst_cell.value = src_cell.value

    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy.copy(src_cell.protection)


def _copy_sheet_to_workbook(src_path: str, dst_wb: Workbook, title: str):
    src_wb = load_workbook(src_path)
    src_ws = src_wb.worksheets[0]

    dst_ws = dst_wb.create_sheet(title)

    for row in src_ws.iter_rows():
        for src_cell in row:
            _copy_cell(src_cell, dst_ws.cell(src_cell.row, src_cell.column))

    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width

    for row_idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_idx].height = dim.height

    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    return dst_ws


def _copy_sheet_area_to_target(src_path: str, dst_ws, start_row: int, start_col: int) -> tuple[int, int]:
    """
    Copy first worksheet from source file into target worksheet at start_row/start_col.
    Used for the left block of Tax Reconciliation sheet.
    """
    src_wb = load_workbook(src_path)
    src_ws = src_wb.worksheets[0]

    for row in src_ws.iter_rows():
        for src_cell in row:
            dst_cell = dst_ws.cell(
                row=start_row + src_cell.row - 1,
                column=start_col + src_cell.column - 1,
            )
            _copy_cell(src_cell, dst_cell)

    # Copy column widths approximately
    for col_idx in range(1, src_ws.max_column + 1):
        src_letter = get_column_letter(col_idx)
        dst_letter = get_column_letter(start_col + col_idx - 1)
        dst_ws.column_dimensions[dst_letter].width = src_ws.column_dimensions[src_letter].width

    last_row = start_row + src_ws.max_row - 1
    last_col = start_col + src_ws.max_column - 1

    return last_row, last_col


def _write_side_labels(
    ws,
    labelled_df: pd.DataFrame,
    source_start_row: int,
    itr_col: int,
    review_col: int,
):
    """
    Write ITR Ref and Review note beside copied raw PL/BS block.
    Only highlight medium/low confidence side cells.
    """
    ws.cell(source_start_row, itr_col, "ITR Ref")
    ws.cell(source_start_row, review_col, "Review note")

    ws.cell(source_start_row, itr_col).font = RED_FONT
    ws.cell(source_start_row, review_col).font = RED_FONT

    for _, row in labelled_df.iterrows():
        source_row = row.get("Source Row")

        if pd.isna(source_row):
            continue

        excel_row = source_start_row + int(source_row) - 1

        itr_ref = row.get("ITR Ref", "")
        review_note = row.get("Review Note", "")
        confidence = str(row.get("Confidence", "")).lower()

        itr_cell = ws.cell(excel_row, itr_col, itr_ref)
        review_cell = ws.cell(excel_row, review_col, review_note)

        if itr_ref:
            itr_cell.font = RED_FONT

        if confidence in {"medium", "low"}:
            itr_cell.fill = REVIEW_FILL
            review_cell.fill = REVIEW_FILL


def _safe(value):
    if pd.isna(value):
        return None
    return value


def _write_table(ws, df: pd.DataFrame, title: str, start_row: int, start_col: int) -> tuple[int, int]:
    if df is None or df.empty:
        ws.cell(start_row, start_col, title)
        return start_row, start_col

    last_col = start_col + len(df.columns) - 1

    for col in range(start_col, last_col + 1):
        cell = ws.cell(start_row, col)
        cell.fill = TITLE_FILL
        cell.font = TITLE_FONT

    ws.cell(start_row, start_col, title)

    header_row = start_row + 1
    for idx, col_name in enumerate(df.columns):
        cell = ws.cell(header_row, start_col + idx, str(col_name))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for r_idx, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        row_text = " ".join(str(x) for x in row.values)

        for c_idx, col_name in enumerate(df.columns):
            value = _safe(row[col_name])
            cell = ws.cell(r_idx, start_col + c_idx, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if isinstance(value, (int, float)):
                cell.number_format = '$#,##0.00;($#,##0.00);-'

            if str(row.get("ITR Ref", "")).strip():
                if col_name == "ITR Ref":
                    cell.font = RED_FONT

        if any(key in row_text for key in ["Accounting Profit", "Taxable Income", "Tax Payable"]):
            for c in range(start_col, last_col + 1):
                ws.cell(r_idx, c).font = BOLD_FONT
                ws.cell(r_idx, c).border = THICK_BORDER

    return header_row + len(df), last_col


def _format_sheet(ws):
    ws.freeze_panes = "A2"

    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 35)


def write_workbook(raw_pl_df, raw_bs_df, workpaper):
    logger.info("Writing workbook to %s", OUTPUT_PATH)

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1 and Sheet 2: exact original copies
    _copy_sheet_to_workbook(PL_RAW_PATH, wb, "Xero PL Raw")
    _copy_sheet_to_workbook(BS_RAW_PATH, wb, "Xero BS Raw")

    # Sheet 3: accountant-style workpaper
    ws = wb.create_sheet("Tax Reconciliation")

    # Left block: PL then BS, both in original format
    pl_start_row = 1
    pl_last_row, pl_last_col = _copy_sheet_area_to_target(PL_RAW_PATH, ws, pl_start_row, 1)

    bs_start_row = pl_last_row + 3
    bs_last_row, bs_last_col = _copy_sheet_area_to_target(BS_RAW_PATH, ws, bs_start_row, 1)

    raw_block_last_col = max(pl_last_col, bs_last_col)

    # Side labels beside the raw block
    itr_col = raw_block_last_col + 2
    review_col = raw_block_last_col + 3

    _write_side_labels(
        ws,
        workpaper.labelled_pl,
        source_start_row=pl_start_row,
        itr_col=itr_col,
        review_col=review_col,
    )

    _write_side_labels(
        ws,
        workpaper.labelled_bs,
        source_start_row=bs_start_row,
        itr_col=itr_col,
        review_col=review_col,
    )

    # Middle block: tax reconciliation
    tax_start_col = review_col + 3
    _write_table(
        ws,
        workpaper.tax_reconciliation,
        "Tax Reconciliation",
        1,
        tax_start_col,
    )

    # Right block: losses + R&D
    support_start_col = tax_start_col + len(workpaper.tax_reconciliation.columns) + 2

    current_row, _ = _write_table(
        ws,
        workpaper.carry_forward_losses,
        "Carry Forward Losses",
        1,
        support_start_col,
    )

    _write_table(
        ws,
        workpaper.rd_breakdown,
        "R&D Breakdown",
        current_row + 3,
        support_start_col,
    )

    _format_sheet(ws)

    Path(OUTPUT_PATH).parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)

    logger.info("Workbook saved: %s", OUTPUT_PATH)